from app.services.reports import _write_html


def _mixed_payload():
    return {"risk": {
        "by_currency": {
            "CNY": {"market_value": 100.0, "pnl": 5.0, "gross_notional": 200.0,
                    "one_day_var_proxy": 3.0, "vega": 1.0, "theta": -0.5,
                    "rho": 0.2, "rho_q": 0.1, "delta_cash": 50.0, "gamma_cash": 2.0,
                    "position_count": 2},
            "USD": {"market_value": 10.0, "pnl": 1.0, "gross_notional": 20.0,
                    "one_day_var_proxy": 0.4, "vega": 0.1, "theta": -0.05,
                    "rho": 0.02, "rho_q": 0.01, "delta_cash": 5.0, "gamma_cash": 0.2,
                    "position_count": 1},
        },
        "shared": {"delta": 3.5, "gamma": 0.0, "delta_proxy": 42.0},
        "totals": None, "mixed_currency": True, "currencies": ["CNY", "USD"],
        "positions": [],
    }}


def _single_payload():
    return {"risk": {
        "by_currency": {"CNY": {"market_value": 100.0, "pnl": 5.0,
                                "gross_notional": 200.0, "one_day_var_proxy": 3.0,
                                "vega": 1.0, "theta": -0.5, "rho": 0.2, "rho_q": 0.1,
                                "delta_cash": 50.0, "gamma_cash": 2.0,
                                "position_count": 2}},
        "shared": {"delta": 3.0, "gamma": 0.0, "delta_proxy": 10.0},
        "totals": {"market_value": 100.0, "pnl": 5.0, "delta_proxy": 10.0,
                   "one_day_var_proxy": 3.0, "delta": 3.0},
        "mixed_currency": False, "currencies": ["CNY"], "positions": [],
    }}


def test_html_mixed_currency_breakdown(tmp_path):
    path = tmp_path / "r.html"
    _write_html(path, "Mixed", _mixed_payload())
    html = path.read_text()
    assert "Mixed currency" in html          # the mixed note replaces top cards
    assert "CNY" in html and "USD" in html    # per-currency sections
    assert "Shared" in html                   # shared section
    assert "42.0000" in html                  # shared delta_proxy rendered


def test_html_single_currency_keeps_top_cards(tmp_path):
    path = tmp_path / "r.html"
    _write_html(path, "Single", _single_payload())
    html = path.read_text()
    assert "Market value" in html             # legacy top cards present
    assert "By currency" in html              # plus the breakdown
    assert "Shared" in html


def test_html_missing_by_currency_does_not_crash(tmp_path):
    path = tmp_path / "r.html"
    _write_html(path, "Empty", {"risk": {"positions": []}})
    assert path.exists()


def test_xlsx_has_by_currency_sheet(tmp_path):
    from openpyxl import load_workbook
    from app.services.reports import _write_xlsx

    path = tmp_path / "r.xlsx"
    _write_xlsx(path, "Mixed", _mixed_payload())
    wb = load_workbook(path)
    assert "By Currency" in wb.sheetnames
    ws = wb["By Currency"]
    rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    assert rows[0] == ("currency", "metric", "value")
    body = rows[1:]
    assert any(r[0] == "CNY" and r[1] == "market_value" and r[2] == 100.0 for r in body)
    assert any(r[0] == "USD" for r in body)
    assert any(r[0] == "(shared)" and r[1] == "delta_proxy" and r[2] == 42.0 for r in body)


def test_xlsx_missing_by_currency_does_not_crash(tmp_path):
    from app.services.reports import _write_xlsx
    path = tmp_path / "r.xlsx"
    _write_xlsx(path, "Empty", {"risk": {"positions": []}})
    assert path.exists()
