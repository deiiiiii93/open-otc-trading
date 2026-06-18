"""The downloadable English import templates stay in lock-step with the adapters.

These tests are the executable form of the "single source of truth" design: the
generated workbook must list exactly the schema's columns, and its own example
rows must parse back through the real adapters as supported positions / pricing
rows. If a column is ever added to the schema but not handled by the adapter (or
vice versa), the round-trip assertions break.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from app.services.import_schema import (
    POSITION_COLUMNS,
    POSITIONS_SHEET_NAME,
    PRICING_COLUMNS,
    PRICING_SHEET_NAME,
)
from app.services.import_templates import (
    INSTRUCTIONS_SHEET_NAME,
    build_positions_template,
    build_pricing_parameters_template,
    positions_template_bytes,
    pricing_parameters_template_bytes,
)
from app.services.market_input_workbooks import read_market_rows_with_diagnostics
from app.services.position_adapter import map_trade_row, read_trade_rows


def _save(workbook, tmp_path, name):
    path = tmp_path / name
    workbook.save(path)
    return path


def test_positions_template_headers_match_schema():
    workbook = build_positions_template()
    assert workbook.sheetnames == [POSITIONS_SHEET_NAME, INSTRUCTIONS_SHEET_NAME]
    sheet = workbook[POSITIONS_SHEET_NAME]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [spec.name for spec in POSITION_COLUMNS]


def test_pricing_template_headers_match_schema():
    workbook = build_pricing_parameters_template()
    assert workbook.sheetnames == [PRICING_SHEET_NAME, INSTRUCTIONS_SHEET_NAME]
    sheet = workbook[PRICING_SHEET_NAME]
    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [spec.name for spec in PRICING_COLUMNS]


def test_instructions_sheet_documents_every_column():
    sheet = build_positions_template()[INSTRUCTIONS_SHEET_NAME]
    documented = {row[0].value for row in sheet.iter_rows(min_row=2)}
    assert documented == {spec.name for spec in POSITION_COLUMNS}


def test_positions_template_examples_round_trip_through_adapter(tmp_path):
    path = _save(build_positions_template(), tmp_path, "positions.xlsx")
    rows = read_trade_rows(path)  # default sheet is the English "Positions"
    assert len(rows) >= 2  # vanilla + snowball examples
    product_types = set()
    for _row_number, row in rows:
        mapping = map_trade_row(row)
        assert mapping.mapping_status == "supported", mapping.mapping_error
        product_types.add(mapping.product_type)
    assert {"EuropeanVanillaOption", "SnowballOption"} <= product_types


def test_pricing_template_examples_round_trip_through_adapter(tmp_path):
    path = _save(build_pricing_parameters_template(), tmp_path, "pricing.xlsx")
    rows, duplicates = read_market_rows_with_diagnostics(path)
    assert not duplicates
    assert rows
    for parsed in rows.values():
        assert parsed["volatility"] is not None
        assert parsed["rate"] is not None
        assert parsed["dividend_yield"] is not None


def test_positions_template_endpoint(client):
    response = client.get("/api/positions/import-template")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "positions_import_template.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [POSITIONS_SHEET_NAME, INSTRUCTIONS_SHEET_NAME]


def test_pricing_template_endpoint(client):
    response = client.get("/api/pricing-parameter-profiles/import-template")
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "pricing_parameters_import_template.xlsx" in response.headers["content-disposition"]
    workbook = load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == [PRICING_SHEET_NAME, INSTRUCTIONS_SHEET_NAME]


def test_template_bytes_are_valid_xlsx():
    for payload in (positions_template_bytes(), pricing_parameters_template_bytes()):
        assert payload[:2] == b"PK"  # xlsx is a zip container
        load_workbook(io.BytesIO(payload))
