from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from fastapi.testclient import TestClient

from app import database
from app.config import Settings
from app.main import create_app
from app.models import MarketDataProfile, PricingParameterProfile, PricingParameterRow
from app.schemas import TrySolveMarketIn, TrySolveQuoteRequestIn, TrySolveRowIn, TrySolveRowOut
from app.services.instruments import ensure_instrument
from app.services.quotes import record_quote
from app.services.try_solve import (
    export_try_solve_workbook,
    _maturity_years,
    _row_to_rfq_draft,
    import_try_solve_workbook,
    resolve_try_solve_market,
    solve_try_solve_row,
    validate_try_solve_row,
)
from app.services.try_solve_registry import (
    PRODUCT_KEYS,
    get_try_solve_catalog,
    registry_by_key,
)


def test_try_solve_catalog_contains_all_workbook_products():
    catalog = get_try_solve_catalog()
    keys = {product["product_key"] for product in catalog["products"]}

    assert keys == {
        "autocall",
        "phoenix",
        "vanilla",
        "vertical_spread",
        "digital",
        "binary_convex",
        "single_sf",
        "double_sf",
        "airbag",
        "airbag_spread",
        "asian",
        "call_put_portfolio",
        "ladder_binary",
        "forward",
        "range_accrual",
        "one_touch",
        "double_no_touch",
        "double_one_touch",
        "knock_out_autocall",
    }
    assert PRODUCT_KEYS == tuple(product["product_key"] for product in catalog["products"])


def test_registry_uses_english_labels_and_excel_aliases():
    vanilla = registry_by_key()["vanilla"]

    assert vanilla.label == "Vanilla"
    assert vanilla.excel_sheet == "vanilla"
    assert vanilla.fields["underlying"].label == "Underlying"
    assert "标的代码" in vanilla.fields["underlying"].excel_aliases
    assert vanilla.quote_fields["premium_rate"].label == "Premium Rate"
    assert vanilla.quote_fields["premium_rate"].excel_header == "期权费率"


def test_try_solve_catalog_exposes_end_date_after_start_date():
    autocall = registry_by_key()["autocall"]

    assert list(autocall.fields).index("end_date") == list(autocall.fields).index("start_date") + 1
    assert "到期日" in autocall.fields["end_date"].excel_aliases


def test_maturity_years_uses_end_date_when_tenor_months_absent():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="autocall",
        fields={
            "underlying": "000852.SH",
            "notional": 1_000_000,
            "start_date": "2026-05-14",
            "end_date": "2026-11-14",
        },
        market={"spot": 100, "volatility": 0.2, "rate": 0.03},
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="annualized_coupon",
            target_label="price",
            target_value=0,
        ),
    )

    assert _maturity_years(row) == 184 / 365


def test_try_solve_catalog_field_mutation_does_not_mutate_registry():
    catalog = get_try_solve_catalog()
    first_field = catalog["products"][0]["fields"][0]
    field_key = first_field["key"]

    first_field["label"] = "Mutated"

    product = registry_by_key()[catalog["products"][0]["product_key"]]
    assert product.fields[field_key].label != "Mutated"


def test_try_solve_catalog_marks_quantark_backed_non_vanilla_products_solver_ready():
    products = registry_by_key()
    ready_keys = {
        key
        for key, product in products.items()
        if any(field.solver_ready for field in product.quote_fields.values())
    }

    assert ready_keys == {
        "autocall",
        "phoenix",
        "vanilla",
        "digital",
        "single_sf",
        "double_sf",
        "asian",
        "forward",
        "range_accrual",
        "one_touch",
        "double_no_touch",
        "double_one_touch",
        "knock_out_autocall",
    }
    assert products["vertical_spread"].initial_solver_state == "schema_captured"
    assert products["binary_convex"].initial_solver_state == "schema_captured"
    assert products["airbag"].initial_solver_state == "schema_captured"
    assert products["airbag_spread"].initial_solver_state == "schema_captured"
    assert products["call_put_portfolio"].initial_solver_state == "schema_captured"
    assert products["ladder_binary"].initial_solver_state == "schema_captured"


def test_validate_try_solve_row_builds_every_quantark_backed_product():
    market = {
        "spot": 100,
        "volatility": 0.2,
        "rate": 0.03,
        "dividend_yield": 0.01,
    }
    for product_key, product in registry_by_key().items():
        quote_field = next(
            (field for field in product.quote_fields.values() if field.solver_ready),
            None,
        )
        if quote_field is None:
            continue
        row = TrySolveRowIn(
            row_id=f"manual-{product_key}",
            product_key=product_key,
            fields={
                "underlying": "000300.SH",
                "notional": 1_000_000,
                "start_date": "2026-05-13",
                "tenor_months": 12,
            },
            market=market,
            quote_request=TrySolveQuoteRequestIn(
                quote_field_key=quote_field.key,
                target_label="price",
                target_value=1,
                lower_bound=quote_field.lower_bound,
                upper_bound=quote_field.upper_bound,
                initial_guess=quote_field.initial_guess,
            ),
        )

        validated = validate_try_solve_row(row)

        assert validated.status == "solver_ready", (
            product_key,
            validated.diagnostics,
        )
        assert validated.quantark_product_type == product.quantark_product_type
        assert validated.engine_name == product.default_engine_name


def _write_try_solve_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "vanilla"
    ws.append(["询价类型", "交易对手", "客户方向", "标的代码", "名义本金", "起始日", "期权类型", "行权价", "期权费率", "询价状态"])
    ws.append(["期权", "招商财富", "买", "000016.SH", 30000000, "2024/7/11", "Call", 1.0, None, None])
    px = wb.create_sheet("phoenix")
    px.append(["询价类型", "交易对手", "客户方向", "标的代码", "名义本金", "起始日", "观察频率", "敲出障碍", "派息收益率", "询价状态"])
    px.append(["期权", "招商财富", "买", "000905.SH", 100000000, "2023/11/7", "1m", 1.03, None, None])
    wb.save(path)


def _make_try_solve_client(tmp_path: Path) -> TestClient:
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'api.sqlite3'}",
        artifact_dir=tmp_path / "artifacts",
    )
    return TestClient(create_app(settings))


def _configure_try_solve_db(tmp_path: Path) -> None:
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'try_solve.sqlite3'}",
        artifact_dir=tmp_path / "artifacts",
    )
    database.configure_database(settings)
    database.init_db()


def test_try_solve_catalog_endpoint(tmp_path: Path):
    client = _make_try_solve_client(tmp_path)

    response = client.get("/api/rfq/try-solve/catalog")

    assert response.status_code == 200
    assert len(response.json()["products"]) == 19


def test_try_solve_import_endpoint_accepts_mixed_workbook(tmp_path: Path):
    client = _make_try_solve_client(tmp_path)
    workbook = tmp_path / "rfq.xlsx"
    _write_try_solve_workbook(workbook)

    with workbook.open("rb") as handle:
        response = client.post(
            "/api/rfq/try-solve/import",
            files={"file": ("rfq.xlsx", handle, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

    assert response.status_code == 200
    assert response.json()["summary"]["total_rows"] == 2


def test_try_solve_import_endpoint_rejects_invalid_workbook(tmp_path: Path):
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'api.sqlite3'}",
        artifact_dir=tmp_path / "artifacts",
    )
    client = TestClient(create_app(settings), raise_server_exceptions=False)

    response = client.post(
        "/api/rfq/try-solve/import",
        files={
            "file": (
                "rfq.xlsx",
                b"not an xlsx workbook",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "workbook" in response.json()["detail"].lower()


def test_try_solve_import_endpoint_rejects_zip_without_workbook(tmp_path: Path):
    settings = Settings(
        database_url=f"sqlite:///{tmp_path / 'api.sqlite3'}",
        artifact_dir=tmp_path / "artifacts",
    )
    client = TestClient(create_app(settings), raise_server_exceptions=False)
    payload = BytesIO()
    with ZipFile(payload, "w") as archive:
        archive.writestr("hello.txt", "not a workbook")

    response = client.post(
        "/api/rfq/try-solve/import",
        files={
            "file": (
                "rfq.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 400
    assert "workbook" in response.json()["detail"].lower()


def test_try_solve_export_endpoint_returns_download_url(tmp_path: Path):
    client = _make_try_solve_client(tmp_path)

    response = client.post(
        "/api/rfq/try-solve/export",
        json={
            "scope": "all",
            "rows": [
                {
                    "row_id": "vanilla:2",
                    "source": "manual",
                    "product_key": "vanilla",
                    "product_label": "Vanilla",
                    "fields": {"underlying": "000016.SH"},
                    "raw_values": {"标的代码": "000016.SH"},
                    "quote_request": {"quote_field_key": "strike", "target_label": "price", "target_value": 8},
                    "status": "solved",
                    "diagnostics": [],
                    "solved_value": 98.5,
                    "model_price": 8.0,
                    "residual": 0.0,
                }
            ],
        },
    )

    assert response.status_code == 200
    body = response.json()
    assert body["url"].endswith(".xlsx")


def test_import_try_solve_workbook_parses_mixed_product_rows(tmp_path: Path):
    workbook = tmp_path / "rfq.xlsx"
    _write_try_solve_workbook(workbook)

    batch = import_try_solve_workbook(workbook)

    assert batch.batch_id
    assert len(batch.rows) == 2
    assert [row.product_key for row in batch.rows] == ["vanilla", "phoenix"]
    assert batch.rows[0].fields["underlying"] == "000016.SH"
    assert batch.rows[0].fields["notional"] == 30000000
    assert batch.rows[0].source_sheet == "vanilla"
    assert batch.rows[0].source_row == 2
    assert batch.rows[1].fields["observation_frequency"] == "1m"
    assert batch.summary["total_rows"] == 2


def test_import_try_solve_workbook_preserves_unknown_columns(tmp_path: Path):
    workbook = tmp_path / "rfq.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "vanilla"
    ws.append(["标的代码", "名义本金", "额外列"])
    ws.append(["000016.SH", 1_000_000, "keep me"])
    wb.save(workbook)

    batch = import_try_solve_workbook(workbook)

    assert batch.rows[0].raw_values["额外列"] == "keep me"
    assert any("Unmapped column: 额外列" in item for item in batch.rows[0].diagnostics)


def test_import_try_solve_workbook_recognizes_quote_headers(tmp_path: Path):
    workbook = tmp_path / "rfq.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "vanilla"
    ws.append(["标的代码", "名义本金", "期权费率"])
    ws.append(["000016.SH", 1_000_000, 0.012])
    wb.save(workbook)

    batch = import_try_solve_workbook(workbook)

    row = batch.rows[0]
    assert row.raw_values["期权费率"] == 0.012
    assert row.quote_request.quote_field_key == "premium_rate"
    assert row.status != "solver_ready"
    assert not any("Unmapped column: 期权费率" in item for item in row.diagnostics)


def test_export_try_solve_workbook_writes_status_and_results(tmp_path: Path):
    output = tmp_path / "try-solve-results.xlsx"
    row = TrySolveRowOut(
        row_id="vanilla:2",
        source="excel",
        product_key="vanilla",
        product_label="Vanilla",
        source_sheet="vanilla",
        source_row=2,
        fields={"underlying": "000016.SH", "notional": 1_000_000, "strike": 100},
        raw_values={"标的代码": "000016.SH", "名义本金": 1_000_000},
        quote_request={"quote_field_key": "strike", "target_label": "price", "target_value": 8},
        status="solved",
        solved_value=98.5,
        model_price=8.0,
        residual=0.0,
        quantark_product_type="EuropeanVanillaOption",
        engine_name="BlackScholesEngine",
    )

    export_try_solve_workbook([row], output)

    wb = load_workbook(output, data_only=True)
    ws = wb["vanilla"]
    headers = [cell.value for cell in ws[1]]
    values = dict(zip(headers, [cell.value for cell in ws[2]]))
    assert values["标的代码"] == "000016.SH"
    assert values["Solve Status"] == "solved"
    assert values["Solved Value"] == 98.5
    assert values["Model Price"] == 8.0
    assert values["QuantArk Product"] == "EuropeanVanillaOption"


def test_export_try_solve_workbook_sanitizes_invalid_sheet_title_chars(
    tmp_path: Path,
):
    output = tmp_path / "try-solve-results.xlsx"
    row = TrySolveRowOut(
        row_id="bad/name:2",
        source="excel",
        product_key="vanilla",
        product_label="Vanilla",
        source_sheet="bad/name[]:*?\\",
        source_row=2,
        fields={},
        raw_values={"标的代码": "000016.SH"},
        quote_request={"quote_field_key": "strike", "target_label": "price", "target_value": 8},
        status="solved",
    )

    export_try_solve_workbook([row], output)

    wb = load_workbook(output, data_only=True)
    assert wb.sheetnames == ["bad_name______"]
    ws = wb["bad_name______"]
    assert ws["A2"].value == "000016.SH"


def test_export_try_solve_workbook_allocates_unique_truncated_sheet_titles(
    tmp_path: Path,
):
    output = tmp_path / "try-solve-results.xlsx"
    prefix = "x" * 31
    rows = [
        TrySolveRowOut(
            row_id="long-1:2",
            source="excel",
            product_key="vanilla",
            product_label="Vanilla",
            source_sheet=f"{prefix}-first",
            source_row=2,
            fields={},
            raw_values={"标的代码": "000016.SH"},
            quote_request={"quote_field_key": "strike", "target_label": "price", "target_value": 8},
            status="solved",
        ),
        TrySolveRowOut(
            row_id="long-2:2",
            source="excel",
            product_key="phoenix",
            product_label="Phoenix",
            source_sheet=f"{prefix}-second",
            source_row=2,
            fields={},
            raw_values={"标的代码": "000905.SH"},
            quote_request={"quote_field_key": "strike", "target_label": "price", "target_value": 8},
            status="solved",
        ),
    ]

    export_try_solve_workbook(rows, output)

    wb = load_workbook(output, data_only=True)
    assert wb.sheetnames == [prefix, f"{'x' * 29}_2"]
    assert all(len(sheet_name) <= 31 for sheet_name in wb.sheetnames)


def test_validate_try_solve_row_blocks_missing_required_field():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "notional": 1_000_000,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike", target_value=8
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "missing_terms"
    assert any("underlying" in item for item in validated.diagnostics)


def test_validate_try_solve_row_blocks_nonpositive_target_value():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1_000_000,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike", target_value=0
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "invalid_target"
    assert any("target.value" in item for item in validated.diagnostics)


def test_row_to_rfq_draft_normalizes_premium_percent_target_label():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1_000_000,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
            "tenor_months": 12,
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="premium %",
            target_value=0.1,
            lower_bound=50,
            upper_bound=150,
            initial_guess=100,
        ),
    )

    product = registry_by_key()["vanilla"]
    draft = _row_to_rfq_draft(row, product)

    assert draft.target.label == "premium"
    assert draft.quantity == 10000.0
    assert draft.target.value == 0.1


def test_row_to_rfq_draft_scales_premium_percent_target_to_notional():
    row = TrySolveRowIn(
        row_id="manual-2",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 2000000,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
            "tenor_months": 12,
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="premium %",
            target_value=1,
            lower_bound=50,
            upper_bound=150,
            initial_guess=100,
        ),
    )

    product = registry_by_key()["vanilla"]
    draft = _row_to_rfq_draft(row, product)

    assert draft.target.label == "premium"
    assert draft.quantity == 20000.0
    assert draft.target.value == 1.0


def test_row_to_rfq_draft_uses_explicit_quantity_for_premium_percent_target():
    row = TrySolveRowIn(
        row_id="manual-3",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1000000,
            "quantity": 500,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
            "tenor_months": 12,
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="premium %",
            target_value=1,
            lower_bound=50,
            upper_bound=150,
            initial_guess=100,
        ),
    )

    product = registry_by_key()["vanilla"]
    draft = _row_to_rfq_draft(row, product)

    assert draft.quantity == 500.0
    assert draft.target.label == "premium"
    assert draft.target.value == 20.0


def test_row_to_rfq_draft_uses_initial_price_for_derived_quantity():
    row = TrySolveRowIn(
        row_id="manual-4",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 8000,
            "initial_price": 80,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
            "tenor_months": 12,
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="premium %",
            target_value=1,
            lower_bound=50,
            upper_bound=150,
            initial_guess=100,
        ),
    )

    product = registry_by_key()["vanilla"]
    draft = _row_to_rfq_draft(row, product)

    assert draft.quantity == 100.0
    assert draft.target.label == "premium"
    assert draft.target.value == 0.8


def test_validate_try_solve_row_blocks_zero_notional():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 0,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike", target_value=8
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "missing_terms"
    assert any("notional" in item for item in validated.diagnostics)


def test_validate_try_solve_row_blocks_unsupported_calendar_override():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1_000_000,
            "strike": 100,
            "option_type": "CALL",
            "start_date": "2024-01-01",
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
            "calendar": "SSE",
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike", target_value=8
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "unsupported_market"
    assert any("calendar" in item for item in validated.diagnostics)


def test_validate_try_solve_row_reports_missing_market():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1_000_000,
            "strike": 1.0,
            "option_type": "CALL",
            "start_date": "2024-01-01",
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike", target_value=0.1
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "missing_market"
    assert any("spot" in item for item in validated.diagnostics)


def test_resolve_try_solve_market_uses_pricing_profile_by_source_trade_id(
    tmp_path: Path,
):
    _configure_try_solve_db(tmp_path)
    with database.SessionLocal() as session:
        profile = PricingParameterProfile(
            name="Close 2026-05-13",
            valuation_date=datetime(2026, 5, 13),
            source_type="xlsx",
            status="completed",
        )
        session.add(profile)
        session.flush()
        # Spot is observation-only — recorded against each row's instrument.
        inst_300 = ensure_instrument(session, "000300.SH", source="pricing_profile", status="draft")
        inst_852 = ensure_instrument(session, "000852.SH", source="pricing_profile", status="draft")
        session.flush()
        record_quote(session, instrument_id=inst_300.id, price=90,
                     as_of=datetime(2026, 5, 13), source="xlsx_import", price_type="mid")
        record_quote(session, instrument_id=inst_852.id, price=102,
                     as_of=datetime(2026, 5, 13), source="xlsx_import", price_type="mid")
        session.add_all(
            [
                PricingParameterRow(
                    profile_id=profile.id,
                    source_trade_id="T-OTHER",
                    symbol="000300.SH",
                    instrument_id=inst_300.id,
                    volatility=0.1,
                    rate=0.01,
                    dividend_yield=0.0,
                ),
                PricingParameterRow(
                    profile_id=profile.id,
                    source_trade_id="T-VANILLA",
                    symbol="000852.SH",
                    instrument_id=inst_852.id,
                    volatility=0.25,
                    rate=0.03,
                    dividend_yield=0.01,
                ),
            ]
        )
        session.flush()

        row = TrySolveRowIn(
            row_id="manual-1",
            product_key="vanilla",
            fields={"source_trade_id": "T-VANILLA", "underlying": "000852.SH"},
            market={"pricing_parameter_profile_id": profile.id},
        )

        market, diagnostics = resolve_try_solve_market(session, row)

        assert diagnostics == []
        assert market.spot == 102
        assert market.volatility == 0.25
        assert market.rate == 0.03
        assert market.dividend_yield == 0.01
        assert market.valuation_date == datetime(2026, 5, 13)


def test_resolve_try_solve_market_uses_market_data_latest_close(tmp_path: Path):
    _configure_try_solve_db(tmp_path)
    with database.SessionLocal() as session:
        profile = MarketDataProfile(
            name="CSI300 daily",
            source="akshare",
            symbol="000300.SH",
            asset_class="index",
            start_date="2026-05-01",
            end_date="2026-05-13",
            valuation_date=datetime(2026, 5, 13),
            data={"latest": {"date": "2026-05-13", "close": 108.5}},
            source_metadata={},
        )
        session.add(profile)
        session.flush()

        row = TrySolveRowIn(
            row_id="manual-1",
            product_key="vanilla",
            fields={"underlying": "000300.SH"},
            market={"market_data_profile_id": profile.id},
        )

        market, diagnostics = resolve_try_solve_market(session, row)

        assert diagnostics == []
        assert market.spot == 108.5
        assert market.valuation_date == datetime(2026, 5, 13)


def test_resolve_try_solve_market_manual_values_override_profile_values(
    tmp_path: Path,
):
    _configure_try_solve_db(tmp_path)
    with database.SessionLocal() as session:
        profile = PricingParameterProfile(
            name="Close 2026-05-13",
            valuation_date=datetime(2026, 5, 13),
            source_type="xlsx",
            status="completed",
        )
        session.add(profile)
        session.flush()
        session.add(
            PricingParameterRow(
                profile_id=profile.id,
                source_trade_id="T-VANILLA",
                symbol="000852.SH",
                volatility=0.25,
                rate=0.03,
                dividend_yield=0.01,
            )
        )
        session.flush()

        row = TrySolveRowIn(
            row_id="manual-1",
            product_key="vanilla",
            fields={"source_trade_id": "T-VANILLA", "underlying": "000852.SH"},
            market={
                "pricing_parameter_profile_id": profile.id,
                "spot": 111,
                "volatility": 0.3,
                "rate": 0.04,
                "dividend_yield": 0.02,
            },
        )

        market, diagnostics = resolve_try_solve_market(session, row)

        assert diagnostics == []
        assert market.spot == 111
        assert market.volatility == 0.3
        assert market.rate == 0.04
        assert market.dividend_yield == 0.02


def test_validate_try_solve_row_marks_unsupported_quote_field():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="airbag",
        fields={"underlying": "000905.SH", "notional": 1_000_000},
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="premium_rate", target_value=0.1
        ),
    )

    validated = validate_try_solve_row(row)

    assert validated.status == "unsupported_quote_field"


def test_solve_try_solve_row_returns_vanilla_quantark_result():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000016.SH",
            "notional": 1,
            "strike": 100,
            "option_type": "CALL",
            "tenor_months": 12,
            "start_date": "2024-01-01",
        },
        market={
            "spot": 100,
            "volatility": 0.2,
            "rate": 0.03,
            "dividend_yield": 0.01,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="price",
            target_value=8,
            lower_bound=50,
            upper_bound=150,
            initial_guess=100,
        ),
    )

    solved = solve_try_solve_row(row)

    assert solved.status == "solved"
    assert solved.solved_value is not None
    assert solved.model_price is not None
    assert solved.executable_terms is not None
    assert solved.quantark_product_type == "EuropeanVanillaOption"


def test_solve_try_solve_row_explains_unbracketed_quote_bounds():
    row = TrySolveRowIn(
        row_id="manual-1",
        product_key="vanilla",
        fields={
            "underlying": "000852.SH",
            "notional": 8900,
            "strike": 8500,
            "option_type": "CALL",
            "tenor_months": 3,
            "start_date": "2026-05-13",
        },
        market={
            "spot": 8741,
            "volatility": 0.25,
            "rate": 0.02,
            "dividend_yield": 0.08,
        },
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="strike",
            target_label="price",
            target_value=0.01,
            lower_bound=8000,
            upper_bound=9000,
            initial_guess=8500,
        ),
    )

    solved = solve_try_solve_row(row)

    assert solved.status == "solve_failed"
    diagnostic = solved.diagnostics[0]
    assert (
        "Target price 0.01 is not between model prices at quote Strike bounds "
        "[8,000, 9,000]"
    ) in diagnostic
    assert "8,000 ->" in diagnostic
    assert "9,000 ->" in diagnostic
    assert "Widen Quote Lower/Upper Bound or adjust Target Value." in diagnostic


def test_flat_contract_for_vanilla_uses_build_product_inputs():
    from app.services.try_solve import _flat_contract_for_row, _pricing_market, _maturity_years

    row = TrySolveRowIn(row_id="r1", product_key="vanilla",
                        fields={"underlying": "000905.SH", "option_type": "CALL"},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="strike", initial_guess=100.0))
    product = registry_by_key()["vanilla"]
    market = _pricing_market(row)
    flat = _flat_contract_for_row(row, product, market, _maturity_years(row),
                                  product.quote_fields["strike"])
    assert flat["initial_price"] == 100.0          # S0 from the reference price
    assert "maturity_years" in flat and "maturity" not in flat
    assert flat["option_type"] == "CALL"


def test_flat_contract_for_autocall_carries_snowball_contract_terms():
    from app.services.try_solve import _flat_contract_for_row, _pricing_market, _maturity_years

    row = TrySolveRowIn(row_id="r2", product_key="autocall",
                        fields={"underlying": "000905.SH", "spot": 100.0,
                                "ko_barrier": 103.0, "ki_barrier": 75.0,
                                "lockup_months": 3, "trade_start_date": "2099-01-05",
                                "observation_frequency": "MONTHLY"},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon", initial_guess=0.15))
    product = registry_by_key()["autocall"]
    market = _pricing_market(row)
    flat = _flat_contract_for_row(row, product, market, _maturity_years(row),
                                  product.quote_fields["annualized_coupon"])
    assert "barrier_config" not in flat            # FLAT, not nested
    for key in ("ko_barrier_pct", "ki_barrier_pct", "lockup_months",
                "trade_start_date", "observation_frequency"):
        assert key in flat, key


def test_build_row_termsheet_synthesizes_complete_vanilla():
    from app.services.try_solve import _build_row_termsheet, _pricing_market, _maturity_years

    row = TrySolveRowIn(row_id="r1", product_key="vanilla",
                        fields={"underlying": "000905.SH", "option_type": "CALL"},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="strike", initial_guess=100.0))
    product = registry_by_key()["vanilla"]
    market = _pricing_market(row)
    kwargs, missing = _build_row_termsheet(row, product, market, _maturity_years(row),
                                           product.quote_fields["strike"])
    assert missing == []
    assert kwargs["maturity"] == _maturity_years(row)   # build_product output uses `maturity`
    assert "maturity_years" not in kwargs               # flat input key consumed
    assert kwargs["option_type"] == "CALL"


def test_build_row_termsheet_reports_precise_gap_when_start_date_absent():
    from app.services.try_solve import _build_row_termsheet, _pricing_market, _maturity_years

    # No start_date -> the snowball cannot synthesize a schedule; build_product
    # surfaces the precise contract gap, never the opaque "KO observation" quad.
    row = TrySolveRowIn(row_id="r2", product_key="autocall",
                        fields={"underlying": "000905.SH", "ko_barrier": 103.0, "ki_barrier": 75.0},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon", initial_guess=0.15))
    product = registry_by_key()["autocall"]
    market = _pricing_market(row)
    kwargs, missing = _build_row_termsheet(row, product, market, _maturity_years(row),
                                           product.quote_fields["annualized_coupon"])
    assert kwargs == {}
    assert missing
    assert any("trade_start_date" in m for m in missing)
    assert not any("KO observation" in m for m in missing)


def test_solve_vanilla_row_matches_direct_build(session):
    from app.services.try_solve import solve_try_solve_row

    row = TrySolveRowIn(row_id="r1", product_key="vanilla",
                        fields={"underlying": "000905.SH", "option_type": "CALL",
                                "notional": 1_000_000, "start_date": "2026-05-13", "tenor_months": 12},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="strike",
                                                             initial_guess=100.0, target_label="price", target_value=8.0))
    out = solve_try_solve_row(row, session)
    assert out.status == "solved", out.diagnostics
    # executable terms are a complete, bookable termsheet (build_product output shape)
    assert out.executable_terms["product_kwargs"]["maturity"]   # `maturity`, not `maturity_years`
    assert "maturity_years" not in out.executable_terms["product_kwargs"]


def test_validate_autocall_row_is_solver_ready_via_build_product(session):
    from app.services.try_solve import validate_try_solve_row

    # A snowball row builds through build_product (real synthesized schedule);
    # validate_quantark_build accepts it -> solver_ready (date-agnostic validate).
    row = TrySolveRowIn(row_id="r2", product_key="autocall",
                        fields={"underlying": "000905.SH", "notional": 1_000_000,
                                "start_date": "2026-05-13", "tenor_months": 12,
                                "ko_barrier": 1.03, "ki_barrier": 0.75},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon",
                                                             initial_guess=0.15, target_label="price", target_value=1.0))
    out = validate_try_solve_row(row, session)
    assert out.status == "solver_ready", out.diagnostics


def test_solve_autocall_row_regenerates_schedule_equal_to_direct_build(session):
    from datetime import date, timedelta
    from app.services.try_solve import solve_try_solve_row
    from app.services.domains.product_builders import build_product

    start = (date.today() + timedelta(days=30)).isoformat()
    row = TrySolveRowIn(row_id="r1", product_key="autocall",
                        fields={"underlying": "000905.SH", "notional": 1_000_000,
                                "start_date": start, "tenor_months": 12,
                                "ko_barrier": 1.03, "ki_barrier": 0.75},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon",
                                                             initial_guess=0.15, target_label="price", target_value=5.0))
    out = solve_try_solve_row(row, session)
    assert out.status == "solved", out.diagnostics
    solved = out.solved_value
    sched = out.executable_terms["product_kwargs"]["barrier_config"]["ko_observation_schedule"]
    # every record's coupon reflects the SOLVED value, not the 0.15 placeholder
    assert all(abs(r["return_rate"] - solved) < 1e-9 for r in sched["records"])
    # and a direct build of the same solved economics yields an identical schedule
    direct = build_product("SnowballOption",
                           {"initial_price": 100.0, "strike": 100.0, "maturity_years": 1.0,
                            "ko_barrier_pct": 103.0, "ki_barrier_pct": 75.0, "ko_rate": solved,
                            "lockup_months": 0, "trade_start_date": start, "observation_frequency": "MONTHLY",
                            "contract_multiplier": 1.0},
                           underlying="000905.SH", currency="USD")
    assert direct.ok
    assert sched == direct.product_kwargs["barrier_config"]["ko_observation_schedule"]


def test_double_touch_flat_contract_sets_touch_type_from_product_key():
    from app.services.try_solve import (
        _flat_contract_for_row, _pricing_market, _maturity_years,
    )

    def _flat(product_key):
        row = TrySolveRowIn(
            row_id="r1", product_key=product_key,
            fields={"underlying": "000905.SH", "notional": 1_000_000,
                    "start_date": "2026-05-13", "tenor_months": 12,
                    "upper_barrier": 1.2, "lower_barrier": 0.8},
            market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
            quote_request=TrySolveQuoteRequestIn(
                quote_field_key="rebate", initial_guess=0.1,
                target_label="price", target_value=0.2),
        )
        product = registry_by_key()[product_key]
        market = _pricing_market(row)
        return _flat_contract_for_row(
            row, product, market, _maturity_years(row),
            product.quote_fields["rebate"],
        )

    no_touch = _flat("double_no_touch")
    one_touch = _flat("double_one_touch")
    assert no_touch["touch_type"] == "DOUBLE_NO_TOUCH"
    assert one_touch["touch_type"] == "DOUBLE_ONE_TOUCH"
    # flat-contract input keys (build_product INPUT names), not QuantArk kwargs
    for flat in (no_touch, one_touch):
        assert {"upper_barrier", "lower_barrier", "cash_payoff",
                "initial_price", "maturity_years"} <= set(flat)


def test_double_one_touch_routes_through_build_product(session):
    from app.services.try_solve import _build_row_termsheet, _pricing_market, _maturity_years

    row = TrySolveRowIn(
        row_id="r1", product_key="double_one_touch",
        fields={"underlying": "000905.SH", "notional": 1_000_000,
                "start_date": "2026-05-13", "tenor_months": 12,
                "upper_barrier": 1.2, "lower_barrier": 0.8},
        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="rebate", initial_guess=0.1,
            target_label="price", target_value=0.2),
    )
    product = registry_by_key()["double_one_touch"]
    market = _pricing_market(row)
    kwargs, missing = _build_row_termsheet(
        row, product, market, _maturity_years(row),
        product.quote_fields["rebate"],
    )
    assert missing == []
    # build_product output: rebate (not cash_payoff), touch direction preserved
    assert kwargs["touch_type"] == "DOUBLE_ONE_TOUCH"
    assert "rebate" in kwargs and "cash_payoff" not in kwargs
    assert set(kwargs) == {"maturity", "upper_barrier", "lower_barrier",
                           "rebate", "touch_type"}


def test_snowball_row_frequency_alias_normalizes_to_canonical():
    from app.services.try_solve import _flat_contract_for_row, _pricing_market, _maturity_years

    def _flat(freq):
        row = TrySolveRowIn(row_id="r1", product_key="autocall",
                            fields={"underlying": "000905.SH", "notional": 1_000_000,
                                    "start_date": "2026-05-13", "tenor_months": 12,
                                    "ko_barrier": 1.03, "ki_barrier": 0.75,
                                    "observation_frequency": freq},
                            market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                            quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon", initial_guess=0.15))
        product = registry_by_key()["autocall"]
        market = _pricing_market(row)
        return _flat_contract_for_row(row, product, market, _maturity_years(row),
                                      product.quote_fields["annualized_coupon"])

    assert _flat("1m")["observation_frequency"] == "MONTHLY"
    assert _flat("quarterly")["observation_frequency"] == "QUARTERLY"
    assert _flat("6M")["observation_frequency"] == "SEMI_ANNUAL"
    assert _flat("MONTHLY")["observation_frequency"] == "MONTHLY"


def test_product_kwargs_for_row_legacy_producer_is_removed():
    import app.services.try_solve as ts
    # The legacy per-row QuantArk-kwargs producer is fully retired; every
    # supported product key now builds via build_product.
    assert not hasattr(ts, "_product_kwargs_for_row")
    assert ts._MIGRATED_PRODUCT_KEYS == ts._SUPPORTED_SOLVE_PRODUCT_KEYS


def test_double_no_touch_routes_through_build_product(session):
    from app.services.try_solve import _build_row_termsheet, _pricing_market, _maturity_years

    row = TrySolveRowIn(
        row_id="r1", product_key="double_no_touch",
        fields={"underlying": "000905.SH", "notional": 1_000_000,
                "start_date": "2026-05-13", "tenor_months": 12,
                "upper_barrier": 1.2, "lower_barrier": 0.8},
        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
        quote_request=TrySolveQuoteRequestIn(
            quote_field_key="rebate", initial_guess=0.1,
            target_label="price", target_value=0.2),
    )
    product = registry_by_key()["double_no_touch"]
    market = _pricing_market(row)
    kwargs, missing = _build_row_termsheet(
        row, product, market, _maturity_years(row),
        product.quote_fields["rebate"],
    )
    assert missing == []
    assert kwargs["touch_type"] == "DOUBLE_NO_TOUCH"
    assert "rebate" in kwargs and "cash_payoff" not in kwargs
    assert set(kwargs) == {"maturity", "upper_barrier", "lower_barrier",
                           "rebate", "touch_type"}


def test_snowball_products_expose_schedule_override_fields():
    products = registry_by_key()
    for key in ("autocall", "phoenix", "knock_out_autocall"):
        keys = set(products[key].fields)
        assert "observation_frequency" in keys, key
        assert "lockup_months" in keys, key
    freq = products["autocall"].fields["observation_frequency"]
    assert freq.field_type == "select"
    assert freq.options == ("MONTHLY", "QUARTERLY", "SEMI_ANNUAL")
    assert products["autocall"].fields["lockup_months"].field_type == "number"


def test_quarterly_override_flows_to_a_quarterly_schedule(session):
    from datetime import date, timedelta
    from app.services.try_solve import solve_try_solve_row

    start = (date.today() + timedelta(days=30)).isoformat()
    row = TrySolveRowIn(row_id="r1", product_key="autocall",
                        fields={"underlying": "000905.SH", "notional": 1_000_000,
                                "start_date": start, "tenor_months": 12,
                                "ko_barrier": 1.03, "ki_barrier": 0.75,
                                "observation_frequency": "QUARTERLY", "lockup_months": 3},
                        market=TrySolveMarketIn(spot=100.0, volatility=0.2, rate=0.03, dividend_yield=0.0),
                        quote_request=TrySolveQuoteRequestIn(quote_field_key="annualized_coupon",
                                                             initial_guess=0.15, target_label="price", target_value=5.0))
    out = solve_try_solve_row(row, session)
    assert out.status == "solved", out.diagnostics
    schedule = out.executable_terms["product_kwargs"]["barrier_config"]["ko_observation_schedule"]
    # quarterly over 1y with a 3-month lockup -> 4 observations (months 3,6,9,12)
    assert len(schedule["records"]) == 4
    assert schedule["frequency"] == "QUARTERLY"
