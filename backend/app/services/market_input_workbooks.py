from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook

from .import_schema import PRICING_REQUIRED_HEADERS, PricingCol
from .position_adapter import normalize_symbol, parse_number, text_value


def read_market_rows_with_diagnostics(
    path: Path,
    *,
    sheet_name: str | None = None,
) -> tuple[dict[str, dict[str, Any]], list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    missing = sorted(PRICING_REQUIRED_HEADERS - set(headers))
    if missing:
        raise ValueError(f"Missing required market headers: {', '.join(missing)}")

    rows: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    for source_row, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        trade_id = text_value(row.get(PricingCol.TRADE_ID))
        if not trade_id:
            continue
        if trade_id in rows and trade_id not in duplicates:
            duplicates.append(trade_id)
        rows[trade_id] = {
            "trade_id": trade_id,
            "symbol": normalize_symbol(row.get(PricingCol.UNDERLYING_CODE)),
            "volatility": parse_number(row.get(PricingCol.VOLATILITY), None),
            "rate": parse_number(row.get(PricingCol.RISK_FREE_RATE), None),
            "dividend_yield": parse_number(row.get(PricingCol.DIVIDEND_BORROW_YIELD), None),
            "spot": parse_number(row.get(PricingCol.UNDERLYING_PRICE), None),
            "source_row": source_row,
            "raw": row,
        }
    return rows, sorted(duplicates)


def infer_valuation_date(path: Path) -> datetime | None:
    text = path.name
    patterns = [
        r"(20\d{2})[-_年](\d{1,2})[-_月](\d{1,2})",
        r"(20\d{2})(\d{2})(\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        year, month, day = (int(part) for part in match.groups())
        try:
            return datetime(year, month, day)
        except ValueError:
            continue
    return None
