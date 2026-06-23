from __future__ import annotations

from datetime import date, datetime
from math import isfinite
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session, selectinload

from ..models import MarketDataProfile, PricingParameterProfile, PricingParameterRow
from ..schemas import (
    PricingEnvironmentSnapshot,
    RFQRequestDraft,
    TrySolveBatchOut,
    TrySolveMarketIn,
    TrySolveQuoteRequestIn,
    TrySolveRowIn,
    TrySolveRowOut,
)
from .position_adapter import normalize_symbol
from .quantark import solve_rfq, validate_quantark_build
from .try_solve_registry import (
    TrySolveField,
    TrySolveProduct,
    TrySolveQuoteField,
    registry_by_key,
    registry_by_sheet,
)


EXPORT_COLUMNS = (
    "Solve Status",
    "Model Price",
    "Residual",
    "Error",
    "Solved Field",
    "Solved Value",
    "Target Label",
    "Target Value",
    "QuantArk Product",
    "Engine",
)
INVALID_SHEET_TITLE_CHARS = frozenset("[]:*?/\\")
MARKET_FIELDS = ("valuation_date", "spot", "volatility", "rate", "dividend_yield")
RAW_TRADE_ID_ALIASES = (
    "source_trade_id",
    "source trade id",
    "trade_id",
    "trade id",
    "position_id",
    "position id",
    "交易编号",
    "合约编号",
)
_TARGET_LABEL_ALIASES = {
    "premium %": "premium",
    "premium%": "premium",
}


def import_try_solve_workbook(path: str | Path) -> TrySolveBatchOut:
    workbook = load_workbook(path, data_only=True)
    products_by_sheet = registry_by_sheet()
    rows: list[TrySolveRowOut] = []
    batch_id = f"try-solve-{uuid4().hex[:12]}"

    for sheet_name in workbook.sheetnames:
        product = products_by_sheet.get(sheet_name)
        if product is None:
            continue

        worksheet = workbook[sheet_name]
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
        )
        headers = [_normalize_header(header) for header in header_row]
        aliases = _alias_map(product.fields)
        quote_aliases = _quote_alias_map(product.quote_fields)

        for row_index, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if not _has_value(values):
                continue

            fields: dict[str, Any] = {}
            raw_values: dict[str, Any] = {}
            diagnostics: list[str] = []
            quote_field_key: str | None = None

            for header, value in zip(headers, values):
                if header == "":
                    continue

                raw_values[header] = value
                field_key = aliases.get(header)
                if field_key is not None:
                    fields[field_key] = value
                elif header in quote_aliases:
                    if value not in (None, "") and quote_field_key is None:
                        quote_field_key = quote_aliases[header]
                elif value not in (None, ""):
                    diagnostics.append(f"Unmapped column: {header}")

            rows.append(
                TrySolveRowOut(
                    row_id=f"{sheet_name}:{row_index}",
                    source="excel",
                    product_key=product.product_key,
                    product_label=product.label,
                    source_sheet=sheet_name,
                    source_row=row_index,
                    fields=fields,
                    raw_values=raw_values,
                    quote_request=TrySolveQuoteRequestIn(
                        quote_field_key=quote_field_key or "premium_rate"
                    ),
                    status="draft",
                    diagnostics=diagnostics,
                    quantark_product_type=product.quantark_product_type,
                    engine_name=product.default_engine_name,
                )
            )

    return TrySolveBatchOut(
        batch_id=batch_id,
        rows=rows,
        summary={
            "total_rows": len(rows),
            "solver_ready": sum(1 for row in rows if row.status == "solver_ready"),
            "schema_captured": sum(1 for row in rows if row.status != "solver_ready"),
        },
    )


def export_try_solve_workbook(
    rows: list[TrySolveRowOut], output_path: str | Path
) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    grouped_rows: dict[str, list[TrySolveRowOut]] = {}
    for row in rows:
        sheet_name = row.source_sheet or row.product_key
        grouped_rows.setdefault(sheet_name, []).append(row)

    if not grouped_rows:
        grouped_rows["try_solve"] = []

    used_sheet_titles: set[str] = set()
    for group_name, group_rows in grouped_rows.items():
        worksheet = workbook.create_sheet(
            title=_allocate_sheet_title(group_name, used_sheet_titles)
        )
        headers = _ordered_raw_headers(group_rows)
        for column in EXPORT_COLUMNS:
            if column not in headers:
                headers.append(column)
        worksheet.append(headers)

        for row in group_rows:
            values = {**row.raw_values, **_export_values(row)}
            worksheet.append([values.get(header) for header in headers])

    workbook.save(path)
    return path


def validate_try_solve_row(
    row: TrySolveRowIn, session: Session | None = None
) -> TrySolveRowOut:
    product = registry_by_key().get(row.product_key)
    if product is None:
        return _row_out(row, "mapping_pending", ["Unknown product key"])

    diagnostics = _required_field_errors(row, product)
    resolved_row, market_diagnostics = _row_with_resolved_market(session, row)
    market_errors = _market_errors(resolved_row)
    if market_diagnostics or market_errors:
        return _row_out(
            resolved_row,
            "missing_market",
            diagnostics + market_diagnostics + market_errors,
            product,
        )

    market_override_errors = _unsupported_market_errors(resolved_row)
    if market_override_errors:
        return _row_out(
            resolved_row,
            "unsupported_market",
            diagnostics + market_override_errors,
            product,
        )

    quote_field = product.quote_fields.get(row.quote_request.quote_field_key)
    if quote_field is None or not quote_field.solver_ready:
        return _row_out(
            resolved_row,
            "unsupported_quote_field",
            diagnostics
            + [f"Unsupported quote field: {row.quote_request.quote_field_key}"],
            product,
        )

    target_errors = _target_errors(row)
    if target_errors:
        return _row_out(
            resolved_row,
            "invalid_target",
            diagnostics + target_errors,
            product,
        )

    term_errors = _required_term_errors(row)
    if diagnostics or term_errors:
        return _row_out(resolved_row, "missing_terms", diagnostics + term_errors, product)

    # Every supported key builds through the single producer: surface precise
    # contract gaps (e.g. a snowball missing schedule inputs), never the opaque
    # quad error.
    market = _pricing_market(resolved_row)
    kwargs, missing = _build_row_termsheet(
        resolved_row, product, market, _maturity_years(resolved_row), quote_field
    )
    if missing:
        return _row_out(resolved_row, "missing_terms", diagnostics + missing, product)
    result = validate_quantark_build(
        product.quantark_product_type, kwargs, market,
        product.default_engine_name, {},
    )
    if not result.ok:
        return _row_out(
            resolved_row,
            "quantark_build_failed",
            diagnostics + [result.error or "QuantArk product build failed"],
            product,
        )

    return _row_out(resolved_row, "solver_ready", diagnostics, product)


def solve_try_solve_row(
    row: TrySolveRowIn, session: Session | None = None
) -> TrySolveRowOut:
    validated = validate_try_solve_row(row, session)
    if validated.status != "solver_ready":
        return validated

    product = registry_by_key().get(row.product_key)
    if product is None:
        return validated

    resolved_row, _diagnostics = _row_with_resolved_market(session, row)
    flat_draft = _row_to_rfq_draft(resolved_row, product)
    market = _pricing_market(resolved_row)
    quote_field = product.quote_fields[row.quote_request.quote_field_key]

    priced_draft = flat_draft
    if row.product_key in _MIGRATED_PRODUCT_KEYS:
        # Synthesize the complete termsheet the solver operates on, but keep the
        # FLAT draft for executable_terms so snowball families regenerate every
        # per-record schedule rate from the solved value (decision 7).
        exec_kwargs, missing = _build_row_termsheet(
            resolved_row, product, market, _maturity_years(resolved_row), quote_field
        )
        if missing:
            return _row_out(resolved_row, "missing_terms", missing, product)
        priced_draft = flat_draft.model_copy(update={"product_kwargs": exec_kwargs})

    result = solve_rfq(priced_draft)
    if not result.ok:
        return _row_out(
            resolved_row,
            "solve_failed",
            [result.error or "QuantArk solve failed"],
            product,
        )

    quote_payload = result.data or {}
    from .rfq import executable_terms_for_quote, _BUILD_PRODUCT_FAMILIES

    # Snowball families regenerate from the FLAT draft (decision 7); simple
    # families top-level-patch the COMPLETE termsheet.
    terms_draft = flat_draft if flat_draft.product_type in _BUILD_PRODUCT_FAMILIES else priced_draft
    executable_terms = executable_terms_for_quote(terms_draft, "solve", quote_payload)
    return _row_out(
        resolved_row,
        "solved",
        validated.diagnostics,
        product,
        solved_value=quote_payload.get("solved_value"),
        model_price=quote_payload.get("achieved_price"),
        residual=quote_payload.get("residual"),
        executable_terms=executable_terms,
    )


def resolve_try_solve_market(
    session: Session | None, row: TrySolveRowIn
) -> tuple[TrySolveMarketIn, list[str]]:
    diagnostics: list[str] = []
    payload: dict[str, Any] = {
        "pricing_parameter_profile_id": row.market.pricing_parameter_profile_id,
        "market_data_profile_id": row.market.market_data_profile_id,
    }

    if row.market.pricing_parameter_profile_id is not None:
        if session is None:
            diagnostics.append("Pricing parameter profile requires a database session")
        else:
            pricing_profile = (
                session.query(PricingParameterProfile)
                .options(selectinload(PricingParameterProfile.rows))
                .filter(PricingParameterProfile.id == row.market.pricing_parameter_profile_id)
                .one_or_none()
            )
            if pricing_profile is None:
                diagnostics.append(
                    f"Pricing parameter profile not found: {row.market.pricing_parameter_profile_id}"
                )
            else:
                pricing_row = _match_pricing_parameter_row(pricing_profile.rows, row)
                if pricing_row is None:
                    diagnostics.append(
                        f"Pricing parameter profile {pricing_profile.id} has no row matching "
                        f"{_market_match_description(row)}"
                    )
                else:
                    payload.update(
                        _market_payload_from_pricing_row(
                            session, pricing_profile, pricing_row
                        )
                    )

    if row.market.market_data_profile_id is not None:
        if session is None:
            diagnostics.append("Market data profile requires a database session")
        else:
            market_profile = session.get(
                MarketDataProfile, row.market.market_data_profile_id
            )
            if market_profile is None:
                diagnostics.append(
                    f"Market data profile not found: {row.market.market_data_profile_id}"
                )
            else:
                spot = _spot_from_market_data_profile(market_profile)
                if spot is None:
                    diagnostics.append(
                        f"Market data profile {market_profile.id} has no spot or latest close"
                    )
                else:
                    payload["spot"] = spot
                    payload.setdefault("valuation_date", market_profile.valuation_date)

    for key in MARKET_FIELDS:
        value = getattr(row.market, key)
        if value is not None:
            payload[key] = value
    for key in ("day_count_convention", "bus_days_in_year", "calendar"):
        value = getattr(row.market, key)
        if value is not None:
            payload[key] = value

    return TrySolveMarketIn(**payload), diagnostics


def _row_out(
    row: TrySolveRowIn,
    status: str,
    diagnostics: list[str],
    product: TrySolveProduct | None = None,
    **updates: Any,
) -> TrySolveRowOut:
    return TrySolveRowOut(
        **row.model_dump(),
        product_label=product.label if product is not None else row.product_key,
        status=status,
        diagnostics=diagnostics,
        quantark_product_type=(
            product.quantark_product_type if product is not None else None
        ),
        engine_name=product.default_engine_name if product is not None else None,
        **updates,
    )


def _row_with_resolved_market(
    session: Session | None, row: TrySolveRowIn
) -> tuple[TrySolveRowIn, list[str]]:
    market, diagnostics = resolve_try_solve_market(session, row)
    return row.model_copy(update={"market": market}), diagnostics


def _match_pricing_parameter_row(
    pricing_rows: list[PricingParameterRow], row: TrySolveRowIn
) -> PricingParameterRow | None:
    ordered_rows = sorted(pricing_rows, key=lambda item: item.source_trade_id or "")
    rows_by_trade_id = {
        _normalize_match_text(pricing_row.source_trade_id): pricing_row
        for pricing_row in ordered_rows
        if _normalize_match_text(pricing_row.source_trade_id)
    }
    for trade_id in _trade_id_candidates(row):
        pricing_row = rows_by_trade_id.get(_normalize_match_text(trade_id))
        if pricing_row is not None:
            return pricing_row

    symbol = _symbol_candidate(row)
    if symbol:
        normalized_symbol = normalize_symbol(symbol)
        for pricing_row in ordered_rows:
            if normalize_symbol(pricing_row.symbol) == normalized_symbol:
                return pricing_row
    return None


def _trade_id_candidates(row: TrySolveRowIn) -> list[str]:
    candidates: list[str] = []
    field_trade_id = _text_value(row.fields.get("source_trade_id"))
    if field_trade_id:
        candidates.append(field_trade_id)

    normalized_raw_values = {
        _normalize_match_text(key): value for key, value in row.raw_values.items()
    }
    for alias in RAW_TRADE_ID_ALIASES:
        raw_value = _text_value(normalized_raw_values.get(_normalize_match_text(alias)))
        if raw_value:
            candidates.append(raw_value)

    seen: set[str] = set()
    unique_candidates: list[str] = []
    for candidate in candidates:
        normalized = _normalize_match_text(candidate)
        if normalized and normalized not in seen:
            seen.add(normalized)
            unique_candidates.append(candidate)
    return unique_candidates


def _symbol_candidate(row: TrySolveRowIn) -> str | None:
    for value in (row.fields.get("underlying"), row.fields.get("symbol")):
        text = _text_value(value)
        if text:
            return text
    normalized_raw_values = {
        _normalize_match_text(key): value for key, value in row.raw_values.items()
    }
    for alias in ("underlying", "symbol", "标的代码", "标的物代码"):
        text = _text_value(normalized_raw_values.get(_normalize_match_text(alias)))
        if text:
            return text
    return None


def _market_match_description(row: TrySolveRowIn) -> str:
    trade_ids = _trade_id_candidates(row)
    if trade_ids:
        return f"source_trade_id={trade_ids[0]}"
    symbol = _symbol_candidate(row)
    if symbol:
        return f"underlying={symbol}"
    return "source_trade_id or underlying"


def _market_payload_from_pricing_row(
    session: Session, profile: PricingParameterProfile, pricing_row: PricingParameterRow
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "valuation_date": profile.valuation_date,
    }
    # r/q/vol live on the row; spot is observation-only (instrument-unification T8):
    # read the latest quote for the row's instrument as of the profile's valuation.
    for key in ("volatility", "rate", "dividend_yield"):
        value = getattr(pricing_row, key)
        if value is not None:
            payload[key] = value
    if pricing_row.instrument_id is not None:
        from .quotes import latest_quote

        quote = latest_quote(
            session, pricing_row.instrument_id, as_of=profile.valuation_date
        )
        if quote is not None:
            payload["spot"] = float(quote.price)
    return payload


def _spot_from_market_data_profile(profile: MarketDataProfile) -> float | None:
    data = profile.data or {}
    for value in (
        data.get("spot"),
        (data.get("latest") or {}).get("close")
        if isinstance(data.get("latest"), dict)
        else None,
    ):
        number = _as_float(value)
        if number is not None:
            return number

    rows = data.get("rows")
    if isinstance(rows, list):
        for item in reversed(rows):
            if not isinstance(item, dict):
                continue
            number = _as_float(item.get("close"))
            if number is not None:
                return number
    return None


def _text_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_match_text(value: Any) -> str:
    text = _text_value(value)
    return text.casefold() if text is not None else ""


def _market_errors(row: TrySolveRowIn) -> list[str]:
    errors: list[str] = []
    positive_fields = {"spot", "volatility"}
    for key in ("spot", "volatility", "rate", "dividend_yield"):
        value = getattr(row.market, key)
        number = _as_float(value)
        if number is None:
            errors.append(f"Missing market {key}")
        elif key in positive_fields and number <= 0:
            errors.append(f"Invalid market {key}")
    return errors


def _unsupported_market_errors(row: TrySolveRowIn) -> list[str]:
    if not _is_blank(row.market.calendar):
        return ["Unsupported market calendar override: calendar"]
    return []


def _target_errors(row: TrySolveRowIn) -> list[str]:
    target = _as_float(row.quote_request.target_value)
    if target is None or target <= 0:
        return ["target.value must be positive for solve mode"]
    return []


def _required_field_errors(
    row: TrySolveRowIn, product: TrySolveProduct
) -> list[str]:
    errors: list[str] = []
    for field in product.fields.values():
        if (
            field.required
            and field.default is None
            and _is_blank(row.fields.get(field.key))
        ):
            errors.append(f"Missing required field: {field.key}")
    return errors


def _required_term_errors(row: TrySolveRowIn) -> list[str]:
    errors: list[str] = []
    if _is_blank(row.fields.get("underlying")):
        errors.append("Missing required term: underlying")
    notional = _as_float(row.fields.get("notional"))
    if notional is None:
        errors.append("Missing required term: notional")
    elif notional <= 0:
        errors.append("notional must be positive")
    quantity = _as_float(row.fields.get("quantity"))
    if quantity is not None and quantity <= 0:
        errors.append("quantity must be positive")
    initial_price = _as_float(row.fields.get("initial_price"))
    if initial_price is not None and initial_price <= 0:
        errors.append("initial_price must be positive")
    if row.product_key not in _SUPPORTED_SOLVE_PRODUCT_KEYS:
        errors.append("Product is not mapped to solver terms")
    if row.product_key in _STRIKE_PRODUCTS and (
        row.quote_request.quote_field_key != "strike"
        and _as_float(row.fields.get("strike")) is None
    ):
        errors.append("Missing required term: strike")
    return errors


def _row_to_rfq_draft(
    row: TrySolveRowIn, product: TrySolveProduct
) -> RFQRequestDraft:
    quote_field = product.quote_fields[row.quote_request.quote_field_key]
    market = _pricing_market(row)
    maturity = _maturity_years(row)

    # Every supported key carries the FLAT term contract (build_product input);
    # the complete termsheet is synthesized downstream (build-through) before solve.
    product_kwargs = _flat_contract_for_row(row, product, market, maturity, quote_field)
    reference = _reference_price(row, market)

    notional = _as_float(row.fields.get("notional"))
    quantity = _rfq_quantity(row, notional, market.spot)
    target_label = _normalize_target_label(row.quote_request.target_label)
    target_value = _as_float(row.quote_request.target_value)
    if target_value is None:
        target_value = 0.0
    if target_label == "premium" and _is_premium_percent_target(row.quote_request.target_label):
        target_value = _premium_percent_target_per_unit(
            target_value,
            notional,
            quantity,
        )

    unknown_values = _unknown_candidate_values(
        quote_field.canonical_path,
        quote_request=row.quote_request,
        quote_field=quote_field,
        reference=reference,
    )

    return RFQRequestDraft(
        client_name=str(row.fields.get("counterparty") or "Demo Client"),
        underlying=str(row.fields.get("underlying") or row.product_key),
        side=_normalize_side(row.fields.get("side")),
        quantity=quantity,
        quote_mode="solve",
        product_type=product.quantark_product_type or "EuropeanVanillaOption",
        product_kwargs=product_kwargs,
        market=market,
        engine_spec={
            "engine_name": product.default_engine_name or "BlackScholesEngine"
        },
        unknown={
            "field_path": quote_field.canonical_path,
            "lower_bound": unknown_values["lower_bound"],
            "upper_bound": unknown_values["upper_bound"],
            "initial_guess": unknown_values["initial_guess"],
            "display_label": quote_field.label,
        },
        target={
            "label": target_label,
            "value": target_value,
        },
    )


_SUPPORTED_SOLVE_PRODUCT_KEYS = frozenset(
    {
        "autocall",
        "phoenix",
        "vanilla",
        "digital",
        "single_sf",
        "double_sf",
        "asian",
        "forward",
        "range_accrual",
        "one_touch",
        "double_no_touch",
        "double_one_touch",
        "knock_out_autocall",
    }
)
_STRIKE_PRODUCTS = frozenset({"vanilla"})


# _MIGRATED_PRODUCT_KEYS == _SUPPORTED_SOLVE_PRODUCT_KEYS: every supported key
# builds through build_product (the single producer). Kept as a named constant
# so call sites remain self-documenting.
_MIGRATED_PRODUCT_KEYS = frozenset({
    "vanilla", "digital", "single_sf", "double_sf", "asian",
    "range_accrual", "one_touch", "forward",
    "autocall", "phoenix", "knock_out_autocall",
    "double_no_touch", "double_one_touch",
})
_SNOWBALL_PRODUCT_KEYS = frozenset({"autocall", "phoenix", "knock_out_autocall"})

# Grid/Excel frequency strings -> build_product's canonical observation tokens
# (FREQUENCY_MONTHS = {MONTHLY:1, QUARTERLY:3, SEMI_ANNUAL:6}).
_FREQUENCY_ALIASES = {
    "1M": "MONTHLY", "M": "MONTHLY", "MONTH": "MONTHLY", "MONTHLY": "MONTHLY",
    "3M": "QUARTERLY", "Q": "QUARTERLY", "QUARTER": "QUARTERLY", "QUARTERLY": "QUARTERLY",
    "6M": "SEMI_ANNUAL", "H": "SEMI_ANNUAL", "SEMIANNUAL": "SEMI_ANNUAL",
    "SEMI_ANNUAL": "SEMI_ANNUAL", "SEMI-ANNUAL": "SEMI_ANNUAL",
}


def _normalize_frequency(value: Any, default: str = "MONTHLY") -> str:
    token = str(value or "").strip().upper()
    return _FREQUENCY_ALIASES.get(token, token or default)


class _UnmigratedProductKey(Exception):
    """Raised by _flat_contract_for_row for product_keys not handled by any
    branch (defensive guard; should be unreachable for registered keys)."""


def _pct_term(row: TrySolveRowIn, field_key: str, reference: float, fallback_mult: float) -> float:
    """A barrier as a percent of initial_price (build_product's *_pct inputs).
    Rows carry absolute / multiple-of-spot barriers; normalize to percent."""
    absolute = _term_price(row, field_key, reference, fallback_mult)
    return (absolute / reference) * 100.0 if reference else absolute


def _flat_contract_for_row(
    row: TrySolveRowIn,
    product: TrySolveProduct,
    market: PricingEnvironmentSnapshot,
    maturity: float,
    quote_field: TrySolveQuoteField,
) -> dict[str, Any]:
    """Map a row to the FLAT term contract that build_product consumes (build
    INPUTS, not QuantArk kwargs). The build-through adapter fills the solve
    target's flat key with its initial guess before build_product (decision 6)."""
    del quote_field
    reference = _reference_price(row, market)
    strike = _term_price(row, "strike", reference, row.quote_request.initial_guess)
    option_type = str(row.fields.get("option_type") or "CALL").upper()
    base: dict[str, Any] = {
        "initial_price": reference,
        "maturity_years": maturity,
        "contract_multiplier": 1.0,
    }

    if row.product_key == "vanilla":
        return {**base, "strike": strike, "option_type": option_type}
    if row.product_key == "digital":
        return {**base, "strike": strike, "option_type": option_type,
                "cash_payoff": _term_amount(row, "payout", reference, 0.1)}
    if row.product_key == "single_sf":
        return {**base, "strike": strike, "option_type": option_type,
                "barrier": _term_price(row, "barrier", reference, 1.2),
                "participation_rate": _term_rate(row, "participation_rate", 1.0)}
    if row.product_key == "double_sf":
        return {**base, "strike": strike, "option_type": option_type,
                "lower_barrier": _term_price(row, "lower_barrier", reference, 0.8),
                "upper_barrier": _term_price(row, "upper_barrier", reference, 1.2),
                "participation_rate": _term_rate(row, "participation_rate", 1.0)}
    if row.product_key == "asian":
        return {**base, "strike": strike, "option_type": option_type,
                "averaging_frequency": str(row.fields.get("averaging_frequency") or "MONTHLY").upper()}
    if row.product_key == "range_accrual":
        return {**base,
                "lower_barrier": _term_price(row, "lower_barrier", reference, 0.8),
                "upper_barrier": _term_price(row, "upper_barrier", reference, 1.2),
                "accrual_rate": _term_rate(row, "coupon_yield", 0.1),
                "observation_frequency": str(row.fields.get("observation_frequency") or "DAILY").upper()}
    if row.product_key == "one_touch":
        barrier = _term_price(row, "barrier", reference, 1.2)
        return {**base, "barrier": barrier,
                "cash_payoff": _term_amount(row, "rebate", reference, 0.1),
                "barrier_direction": "UP" if barrier >= reference else "DOWN",
                "touch_type": "ONE_TOUCH"}
    if row.product_key == "forward":
        return {**base, "underlying": str(row.fields.get("underlying") or row.product_key),
                "basis": _term_rate(row, "basis", 0.0)}
    if row.product_key in {"double_no_touch", "double_one_touch"}:
        # Both keys share the QuantArk class DoubleOneTouchOption; touch direction
        # rides in touch_type, derived from product_key. cash_payoff -> rebate is
        # handled by the builder (consistent with one_touch/digital).
        return {**base,
                "upper_barrier": _term_price(row, "upper_barrier", reference, 1.2),
                "lower_barrier": _term_price(row, "lower_barrier", reference, 0.8),
                "cash_payoff": _term_amount(row, "rebate", reference, 0.1),
                "touch_type": ("DOUBLE_NO_TOUCH"
                               if row.product_key == "double_no_touch"
                               else "DOUBLE_ONE_TOUCH")}
    if row.product_key in _SNOWBALL_PRODUCT_KEYS:
        # try-solve rows capture start_date + barriers; lockup/frequency are not
        # yet grid columns (frontend follow-on), so default to no lockup + monthly
        # observations for the quote estimate (both overridable via row.fields).
        lockup = _as_float(row.fields.get("lockup_months"))
        flat: dict[str, Any] = {
            **base,
            "strike": _term_price(row, "strike", reference, 1.0),
            "ko_barrier_pct": _pct_term(row, "ko_barrier", reference, 1.03),
            "ki_barrier_pct": _pct_term(row, "ki_barrier", reference, 0.75),
            "ko_rate": _term_rate(row, "annualized_coupon", 0.1),
            "lockup_months": lockup if lockup is not None else 0.0,
            "trade_start_date": row.fields.get("start_date"),
            "observation_frequency": _normalize_frequency(row.fields.get("observation_frequency")),
        }
        if row.product_key == "phoenix":
            flat["ko_rate"] = _term_rate(row, "annualized_coupon", 0.0)
            flat["coupon_barrier_pct"] = _pct_term(row, "coupon_barrier", reference, 0.85)
            flat["coupon_rate"] = _term_rate(row, "coupon_yield", 0.1)
        if row.product_key == "knock_out_autocall":
            flat["post_ko_barrier_pct"] = flat["ko_barrier_pct"]
            flat["post_ko_rate"] = flat["ko_rate"]
        return flat
    raise _UnmigratedProductKey(row.product_key)


def _build_row_termsheet(
    row: TrySolveRowIn,
    product: TrySolveProduct,
    market: PricingEnvironmentSnapshot,
    maturity: float,
    quote_field: TrySolveQuoteField,
) -> tuple[dict[str, Any], list[str]]:
    """Complete QuantArk termsheet for a row via build_product (the single
    producer). Every supported product key is migrated, so there is no legacy
    fallback; an unknown family raises _UnmigratedProductKey from
    _flat_contract_for_row. `solve_target` exempts the solved field from
    `missing` (filter_solved). Returns (product_kwargs, missing); missing is
    non-empty iff the row's contract is unfilled (e.g. a snowball missing
    schedule inputs)."""
    from .domains.product_builders import build_product

    flat = _flat_contract_for_row(row, product, market, maturity, quote_field)
    built = build_product(
        product.quantark_product_type or "EuropeanVanillaOption",
        flat,
        underlying=str(row.fields.get("underlying") or row.product_key),
        currency=market.currency,
        solve_target=quote_field.canonical_path,
    )
    if built.missing:
        return {}, built.missing
    return dict(built.product_kwargs), []


def _pricing_market(row: TrySolveRowIn) -> PricingEnvironmentSnapshot:
    payload: dict[str, Any] = {
        "spot": row.market.spot,
        "volatility": row.market.volatility,
        "rate": row.market.rate,
        "dividend_yield": row.market.dividend_yield,
    }
    for key in ("valuation_date", "day_count_convention", "bus_days_in_year"):
        value = getattr(row.market, key)
        if value is not None:
            payload[key] = value
    return PricingEnvironmentSnapshot(**payload)


def _reference_price(row: TrySolveRowIn, market: PricingEnvironmentSnapshot) -> float:
    initial_price = _as_float(row.fields.get("initial_price"))
    if initial_price is not None and initial_price > 0:
        return initial_price
    return float(market.spot)


def _term_price(
    row: TrySolveRowIn,
    field_key: str,
    reference: float,
    fallback: float | None,
) -> float:
    value = _as_float(row.fields.get(field_key))
    if value is None:
        value = fallback
    if value is None:
        value = reference
    return _scale_moneyness(value, reference)


def _term_amount(
    row: TrySolveRowIn,
    field_key: str,
    reference: float,
    fallback_rate: float,
) -> float:
    value = _as_float(row.fields.get(field_key))
    if value is None:
        value = fallback_rate
    return _scale_moneyness(value, reference)


def _term_rate(row: TrySolveRowIn, field_key: str, fallback: float) -> float:
    value = _as_float(row.fields.get(field_key))
    return float(value if value is not None else fallback)


def _scale_moneyness(value: float, reference: float) -> float:
    if reference > 10 and 0 < abs(value) <= 10:
        return float(value * reference)
    return float(value)


def _unknown_candidate_values(
    field_path: str,
    *,
    quote_request: TrySolveQuoteRequestIn,
    quote_field: TrySolveQuoteField,
    reference: float,
) -> dict[str, float | None]:
    """Resolve the solver's [lower, upper, guess] in absolute space.

    Source-aware: a *user-supplied* bound/guess (from ``quote_request``) is taken
    in the units the explicit ``quote_value_mode`` declares — ``percentage`` =>
    ``value/100 * spot``, otherwise absolute as typed. A *registry default* (from
    ``quote_field``, omitted by the user) is stored as moneyness and is scaled to
    the same spot-scaled absolute space the solved term value uses
    (``_term_price`` -> ``_scale_moneyness``), so a price-like field's default
    bounds line up with an absolute user guess.
    """
    price_like = _is_price_like_unknown(field_path)
    percentage = (
        quote_request.quote_value_mode == "percentage" and price_like and reference > 0
    )
    user = {
        "lower_bound": quote_request.lower_bound,
        "upper_bound": quote_request.upper_bound,
        "initial_guess": quote_request.initial_guess,
    }
    default = {
        "lower_bound": quote_field.lower_bound,
        "upper_bound": quote_field.upper_bound,
        "initial_guess": quote_field.initial_guess,
    }
    resolved: dict[str, float | None] = {}
    for key, user_value in user.items():
        if user_value is not None:
            resolved[key] = (
                float(user_value) / 100.0 * reference
                if percentage
                else float(user_value)
            )
        else:
            default_value = default[key]
            resolved[key] = (
                None
                if default_value is None
                else (
                    _scale_moneyness(float(default_value), reference)
                    if price_like
                    else float(default_value)
                )
            )
    return resolved


def _is_price_like_unknown(field_path: str) -> bool:
    return field_path in {
        "strike",
        "barrier",
        "upper_barrier",
        "lower_barrier",
        "barrier_config.ko_barrier",
        "barrier_config.ki_barrier",
        "coupon_config.coupon_barrier",
        "range_config.upper_barrier",
        "range_config.lower_barrier",
    }


def _maturity_years(row: TrySolveRowIn) -> float:
    months = _as_float(row.fields.get("tenor_months"))
    if months is not None and months > 0:
        return months / 12.0

    start = _date_value(row.fields.get("start_date"))
    end = _date_value(row.fields.get("end_date"))
    if start is not None and end is not None and end > start:
        return (end - start).days / 365.0

    days = _as_float(row.fields.get("tenor_days"))
    if days is not None and days > 0:
        return days / 365.0

    return 1.0


def _date_value(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return date.fromisoformat(text[:10])
        except ValueError:
            return None
    return None


def _normalize_side(value: Any) -> str:
    text = str(value or "buy").strip().lower()
    if text in {"sell", "sold", "short", "卖", "卖出"}:
        return "sell"
    return "buy"


def _normalize_target_label(label: str) -> str:
    normalized = str(label or "").strip()
    normalized_lower = normalized.lower()
    if normalized_lower in _TARGET_LABEL_ALIASES:
        return _TARGET_LABEL_ALIASES[normalized_lower]
    return _TARGET_LABEL_ALIASES.get(normalized, normalized)


def _is_premium_percent_target(label: str) -> bool:
    normalized = str(label or "").strip().lower()
    return normalized in {"premium %", "premium%"}


def _rfq_quantity(
    row: TrySolveRowIn,
    notional: float | None,
    fallback_initial_price: float | None,
) -> float:
    quantity = _as_float(row.fields.get("quantity"))
    if quantity is not None and quantity > 0:
        return quantity

    initial_price = _as_float(row.fields.get("initial_price"))
    if initial_price is None:
        initial_price = fallback_initial_price
    if notional is not None and initial_price is not None and initial_price > 0:
        return float(notional / initial_price)
    return float(notional if notional is not None else 1.0)


def _premium_percent_target_per_unit(
    target_percent: float,
    notional: float | None,
    quantity: float,
) -> float:
    total_premium = target_percent * (notional if notional is not None else quantity) / 100.0
    return total_premium / quantity if quantity > 0 else total_premium


def _as_float(value: Any) -> float | None:
    if value is None or value == "" or isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not isfinite(number):
        return None
    return number


def _is_blank(value: Any) -> bool:
    return value is None or value == ""


def _alias_map(fields: dict[str, TrySolveField]) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for field in fields.values():
        aliases[field.label] = field.key
        for alias in field.excel_aliases:
            aliases[alias] = field.key
    return aliases


def _quote_alias_map(fields: dict[str, TrySolveQuoteField]) -> dict[str, str]:
    return {field.excel_header: field.key for field in fields.values()}


def _has_value(values: tuple[Any, ...]) -> bool:
    return any(value not in (None, "") for value in values)


def _ordered_raw_headers(rows: list[TrySolveRowOut]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for header in row.raw_values:
            if header not in seen:
                seen.add(header)
                headers.append(header)
    return headers


def _export_values(row: TrySolveRowOut) -> dict[str, Any]:
    return {
        "Solve Status": row.status,
        "Model Price": row.model_price,
        "Residual": row.residual,
        "Error": "; ".join(row.diagnostics),
        "Solved Field": row.quote_request.quote_field_key,
        "Solved Value": row.solved_value,
        "Target Label": row.quote_request.target_label,
        "Target Value": row.quote_request.target_value,
        "QuantArk Product": row.quantark_product_type,
        "Engine": row.engine_name,
    }


def _allocate_sheet_title(name: str, used_titles: set[str]) -> str:
    base_title = _safe_sheet_title(name)
    title = base_title
    suffix_index = 2
    while title.casefold() in used_titles:
        suffix = f"_{suffix_index}"
        prefix_length = max(0, 31 - len(suffix))
        title = f"{base_title[:prefix_length]}{suffix}"[-31:]
        suffix_index += 1

    used_titles.add(title.casefold())
    return title


def _safe_sheet_title(name: str) -> str:
    title = "".join(
        "_" if char in INVALID_SHEET_TITLE_CHARS else char for char in name
    ).strip()
    if not title:
        title = "try_solve"
    title = title[:31]
    return title or "try_solve"


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
