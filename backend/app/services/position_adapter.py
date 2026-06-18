from __future__ import annotations

from dataclasses import dataclass, replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import Portfolio, Position, PositionImportBatch
from .currency_codes import ISO_4217_CODES, normalize_currency
from .domains.booking import (
    BookingRequest,
    ProductBookingSpec,
    book_position,
    prepare_booking_product_spec,
    set_position_currency,
)
from .domains.schedules import is_china_sse_business_day as _shared_is_business_day
from .domains.position_terms import (
    refresh_position_barrier_state,
    reset_position_term_rows,
    upsert_position_term_rows,
)
from .domains.products import (
    create_or_get_product,
    hydrate_position_product_fields,
    product_spec_from_position_payload,
)
from .underlyings import link_position_underlying
from .import_schema import (
    CLOSED_STATUSES,
    CUSTOM_EUROPEAN,
    CUSTOM_NO_KNOCK_IN,
    PARTIAL_PROTECTION_TAG,
    PHOENIX_TAG,
    POSITION_REQUIRED_HEADERS,
    POSITIONS_SHEET_NAME,
    Direction,
    OptionType,
    PositionCol,
    TradeStatus,
    YesNo,
    STRUCTURE_TYPE_BARRIER_KNOCK_IN,
    STRUCTURE_TYPE_DIGITAL,
    STRUCTURE_TYPE_DOUBLE_SHARKFIN,
    STRUCTURE_TYPE_PHOENIX,
    STRUCTURE_TYPE_PHOENIX_PARTIAL,
    STRUCTURE_TYPE_SINGLE_SHARKFIN,
    STRUCTURE_TYPE_SNOWBALL,
    STRUCTURE_TYPE_SNOWBALL_PARTIAL,
    STRUCTURE_TYPE_VANILLA_AMERICAN,
    STRUCTURE_TYPE_VANILLA_EUROPEAN,
)


TRADE_SHEET = POSITIONS_SHEET_NAME
SUPPORTED_STATUS = "supported"
UNSUPPORTED_STATUS = "unsupported"
ERROR_STATUS = "error"
REQUIRED_HEADERS = POSITION_REQUIRED_HEADERS


@dataclass(frozen=True)
class PositionMapping:
    underlying: str
    product_type: str
    product_kwargs: dict[str, Any]
    engine_name: str
    engine_kwargs: dict[str, Any]
    quantity: float
    entry_price: float
    status: str
    mapping_status: str
    mapping_error: str | None = None
    currency: str | None = None


def import_positions_from_xlsx(
    session: Session,
    *,
    portfolio_id: int,
    xlsx_path: str | Path,
    sheet_name: str = TRADE_SHEET,
) -> PositionImportBatch:
    portfolio = session.get(Portfolio, portfolio_id)
    if portfolio is None:
        raise ValueError(f"Portfolio not found: {portfolio_id}")

    source_path = Path(xlsx_path)
    rows = read_trade_rows(source_path, sheet_name=sheet_name)
    imported_count = 0
    supported_count = 0
    unsupported_count = 0
    error_count = 0
    errors: list[dict[str, Any]] = []
    touched_positions: list[Position] = []

    for row_number, row in rows:
        trade_id = str(row.get(PositionCol.TRADE_ID) or "").strip()
        if not trade_id:
            error_count += 1
            errors.append({"row": row_number, "error": "Missing Trade ID"})
            continue

        source_payload = {"row_number": row_number, "row": make_json_safe(row)}
        try:
            mapping = map_trade_row(row)
        except Exception as exc:
            mapping = _error_mapping(row, str(exc))
            error_count += 1
            errors.append({"row": row_number, "trade_id": trade_id, "error": str(exc)})

        existing = (
            session.query(Position)
            .filter(Position.portfolio_id == portfolio.id, Position.source_trade_id == trade_id)
            .one_or_none()
        )
        try:
            if existing is None:
                position = book_position(
                    session,
                    _booking_request_from_mapping(
                        portfolio_id=portfolio.id,
                        mapping=mapping,
                        source_payload=source_payload,
                        trade_id=trade_id,
                        row_number=row_number,
                    ),
                )
            else:
                position = existing
                # Route updates through the same single gate as inserts so a
                # re-imported phoenix/scalar is validated, not trusted as-is.
                spec = prepare_booking_product_spec(
                    _product_booking_spec_from_mapping(mapping, source_payload),
                    engine_name=mapping.engine_name,
                )
                product = create_or_get_product(session, spec, reuse=True)
                position.product = product
                position.product_id = product.id
                hydrate_position_product_fields(position)
                link_position_underlying(session, position, source="import")
                set_position_currency(position)
                position.engine_name = mapping.engine_name
                position.engine_kwargs = mapping.engine_kwargs
                position.quantity = mapping.quantity
                position.entry_price = mapping.entry_price
                position.status = mapping.status
                position.position_kind = "otc"
                position.source_trade_id = trade_id
                position.source_row = row_number
                position.mapping_status = mapping.mapping_status
                position.mapping_error = mapping.mapping_error
                position.source_payload = source_payload
                reset_position_term_rows(session, position.id)
                session.add(position)
        except ValueError as exc:
            # The single gate rejected this row's terms. Isolate it as an error row
            # instead of aborting the import. For a NEW trade, book an error row (the
            # error mapping's structure-label product_type is not a gated family, so
            # the re-book's gate is a no-op). For an EXISTING trade, demote the
            # persisted position in place — never insert a duplicate, since the loop's
            # source_trade_id lookup is .one_or_none().
            error_count += 1
            errors.append({"row": row_number, "trade_id": trade_id, "error": str(exc)})
            mapping = _error_mapping(row, str(exc))
            if existing is None:
                position = book_position(
                    session,
                    _booking_request_from_mapping(
                        portfolio_id=portfolio.id,
                        mapping=mapping,
                        source_payload=source_payload,
                        trade_id=trade_id,
                        row_number=row_number,
                    ),
                )
            else:
                # Keep the previously-valid product; flag the row as error so the
                # bad re-import is visible without losing the existing booking.
                position = existing
                position.mapping_status = mapping.mapping_status
                position.mapping_error = mapping.mapping_error
                position.position_kind = "otc"
                position.source_row = row_number
                position.source_payload = source_payload
                session.add(position)
        touched_positions.append(position)

        imported_count += 1
        if mapping.mapping_status == SUPPORTED_STATUS:
            supported_count += 1
        elif mapping.mapping_status == UNSUPPORTED_STATUS:
            unsupported_count += 1

    batch = PositionImportBatch(
        portfolio_id=portfolio.id,
        source_path=str(source_path),
        source_sheet=sheet_name,
        row_count=len(rows),
        imported_count=imported_count,
        supported_count=supported_count,
        unsupported_count=unsupported_count,
        error_count=error_count,
        status="completed" if error_count == 0 else "completed_with_errors",
        summary={"errors": errors[:50]},
    )
    session.add(batch)
    session.flush()
    for position in touched_positions:
        upsert_position_term_rows(session, position)
    session.flush()
    for position in touched_positions:
        refresh_position_barrier_state(session, position_id=position.id)
    return batch


def _booking_request_from_mapping(
    *,
    portfolio_id: int,
    mapping: PositionMapping,
    source_payload: dict[str, Any],
    trade_id: str,
    row_number: int,
) -> BookingRequest:
    """Build the import BookingRequest for a mapped row. Single source of truth so
    the insert path and the error-isolation re-book path cannot drift."""
    return BookingRequest(
        portfolio_id=portfolio_id,
        product=_product_booking_spec_from_mapping(mapping, source_payload),
        quantity=mapping.quantity,
        entry_price=mapping.entry_price,
        status=mapping.status,
        source_trade_id=trade_id,
        source_row=row_number,
        mapping_status=mapping.mapping_status,
        mapping_error=mapping.mapping_error,
        source_payload=source_payload,
        engine_name=mapping.engine_name,
        engine_kwargs=mapping.engine_kwargs,
        actor="desk_user",
        source="import",
    )


def _product_booking_spec_from_mapping(
    mapping: PositionMapping,
    source_payload: dict[str, Any],
) -> ProductBookingSpec:
    spec = product_spec_from_position_payload(
        {
            "underlying": mapping.underlying,
            "product_type": mapping.product_type,
            "product_kwargs": mapping.product_kwargs,
            "source_payload": source_payload,
        }
    )
    # Import channel: Currency column wins, else CNY. Never inherit the generic
    # spec default (USD) — Chinese OTC trade sheets are CNY-denominated.
    return ProductBookingSpec(**{**spec.__dict__, "currency": mapping.currency or "CNY"})


def read_trade_rows(path: Path, *, sheet_name: str = TRADE_SHEET) -> list[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    worksheet = workbook[sheet_name]
    header_values = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in header_values]
    missing = sorted(REQUIRED_HEADERS - set(headers))
    if missing:
        raise ValueError(f"Missing required headers: {', '.join(missing)}")

    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row = {headers[index]: value for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if any(value not in (None, "") for value in row.values()):
            rows.append((row_number, row))
    return rows


def map_trade_row(row: dict[str, Any]) -> PositionMapping:
    structure_type = text_value(row.get(PositionCol.STRUCTURE_TYPE))
    mapping_by_structure = {
        STRUCTURE_TYPE_VANILLA_EUROPEAN: _map_european_vanilla,
        STRUCTURE_TYPE_VANILLA_AMERICAN: _map_american_vanilla,
        STRUCTURE_TYPE_DIGITAL: _map_digital,
        STRUCTURE_TYPE_BARRIER_KNOCK_IN: _map_barrier_knock_in,
        STRUCTURE_TYPE_SINGLE_SHARKFIN: _map_single_sharkfin,
        STRUCTURE_TYPE_DOUBLE_SHARKFIN: _map_double_sharkfin,
        STRUCTURE_TYPE_SNOWBALL: _map_snowball,
        STRUCTURE_TYPE_SNOWBALL_PARTIAL: _map_snowball,
        STRUCTURE_TYPE_PHOENIX: _map_phoenix,
        STRUCTURE_TYPE_PHOENIX_PARTIAL: _map_phoenix,
    }
    mapper = mapping_by_structure.get(structure_type)
    if mapper is None:
        return _unsupported_mapping(row, f"Unsupported structure type: {structure_type or '<blank>'}")
    return _apply_row_currency(mapper(row), row)


def _apply_row_currency(mapping: PositionMapping, row: dict[str, Any]) -> PositionMapping:
    """Optional Currency column: blank -> leave None (channel default
    CNY applies downstream); invalid -> raise so the import loop isolates the row
    as an error row, consistent with booking-gate rejections."""
    raw = text_value(row.get(PositionCol.CURRENCY))
    if not raw:
        return mapping
    code = normalize_currency(raw)
    if code not in ISO_4217_CODES:
        raise ValueError(f"Invalid currency code: {raw!r}")
    return replace(mapping, currency=code)


def normalize_symbol(value: Any) -> str:
    text = text_value(value)
    if " - " in text:
        text = text.split(" - ", 1)[0]
    return text.strip().upper()


def text_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_number(value: Any, default: float | None = None) -> float | None:
    if value in (None, ""):
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return default
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1]
    try:
        parsed = float(text)
    except ValueError:
        return default
    return parsed / 100.0 if is_percent else parsed


def parse_number_list(value: Any, default: list[float] | None = None) -> list[float]:
    if value in (None, ""):
        return default or []
    if isinstance(value, (int, float)):
        return [float(value)]
    return [item for item in (parse_number(part) for part in str(value).split(",")) if item is not None]


def parse_date(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        return None


def parse_date_list(value: Any) -> list[datetime]:
    if value in (None, ""):
        return []
    if isinstance(value, datetime):
        return [value]
    dates = []
    for part in str(value).split(","):
        parsed = parse_date(part.strip())
        if parsed is not None:
            dates.append(parsed)
    return dates


def make_json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, dict):
        return {str(key): make_json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [make_json_safe(item) for item in value]
    return value


def _map_european_vanilla(row: dict[str, Any]) -> PositionMapping:
    return _supported_mapping(
        row,
        product_type="EuropeanVanillaOption",
        product_kwargs={
            "strike": _required_number(row, PositionCol.STRIKE_PRICE),
            "option_type": _option_type(row),
            **_vanilla_date_kwargs(row),
        },
        engine_name="BlackScholesEngine",
    )


def _map_american_vanilla(row: dict[str, Any]) -> PositionMapping:
    return _supported_mapping(
        row,
        product_type="AmericanOption",
        product_kwargs={
            "strike": _required_number(row, PositionCol.STRIKE_PRICE),
            "option_type": _option_type(row),
            **_vanilla_date_kwargs(row),
        },
        engine_name="AmericanOptionAnalyticalEngine",
    )


def _map_digital(row: dict[str, Any]) -> PositionMapping:
    initial_price = _required_number(row, PositionCol.INITIAL_PRICE)
    payout_rate = parse_number(row.get(PositionCol.COUPON_RATE), 0.0) or 0.0
    return _supported_mapping(
        row,
        product_type="CashOrNothingDigitalOption",
        product_kwargs={
            "strike": _required_number(row, PositionCol.STRIKE_PRICE),
            "payout": initial_price * payout_rate,
            "option_type": _option_type(row),
            **_vanilla_date_kwargs(row),
        },
        engine_name="DigitalOptionAnalyticalEngine",
    )


def _map_barrier_knock_in(row: dict[str, Any]) -> PositionMapping:
    initial_price = _required_number(row, PositionCol.INITIAL_PRICE)
    barrier = _required_number(row, PositionCol.KNOCK_IN_BARRIER)
    barrier_type = "UP_IN" if barrier > initial_price else "DOWN_IN"
    return _supported_mapping(
        row,
        product_type="BarrierOption",
        product_kwargs={
            "strike": _required_number(row, PositionCol.STRIKE_PRICE),
            "option_type": _option_type(row),
            "barrier": barrier,
            "barrier_type": barrier_type,
            "rebate": initial_price * (parse_number(row.get(PositionCol.NO_KNOCK_IN_COUPON), 0.0) or 0.0),
            "participation_rate": parse_number(row.get(PositionCol.PARTICIPATION_RATE), 1.0) or 1.0,
            "observation_type": "CONTINUOUS",
            **_vanilla_date_kwargs(row),
        },
        engine_name="BarrierAnalyticalEngine",
    )


def _map_single_sharkfin(row: dict[str, Any]) -> PositionMapping:
    return _supported_mapping(
        row,
        product_type="SingleSharkfinOption",
        product_kwargs={
            **_sharkfin_core_kwargs(row),
            "barrier": _single_sharkfin_barrier(row),
            **_sharkfin_observation_kwargs(row),
        },
        engine_name="SingleSharkfinOptionAnalyticalEngine",
    )


def _map_double_sharkfin(row: dict[str, Any]) -> PositionMapping:
    lower_barrier, upper_barrier = _double_sharkfin_barriers(row)
    return _supported_mapping(
        row,
        product_type="DoubleSharkfinOption",
        product_kwargs={
            **_sharkfin_core_kwargs(row),
            "lower_barrier": lower_barrier,
            "upper_barrier": upper_barrier,
            **_double_sharkfin_observation_kwargs(row, lower_barrier, upper_barrier),
        },
        engine_name="DoubleSharkfinOptionAnalyticalEngine",
    )


def _map_snowball(row: dict[str, Any]) -> PositionMapping:
    structure_type = text_value(row.get(PositionCol.STRUCTURE_TYPE))
    return _supported_mapping(
        row,
        product_type="SnowballOption",
        product_kwargs={
            **_autocallable_core_kwargs(row),
            "barrier_config": _snowball_barrier_config(row),
            "payoff_config": _autocallable_payoff_config(row, partial_protection=PARTIAL_PROTECTION_TAG in structure_type),
            "accrual_config": _autocallable_accrual_config(row),
            "_otc_ki_observation_convention": _ki_observation_convention(row),
            "_otc_lifecycle_knocked_in": _autocallable_knocked_in(row),
            "_otc_lifecycle_state": text_value(row.get(PositionCol.TRADE_STATUS)),
        },
        engine_name="SnowballQuadEngine",
        engine_kwargs=_quad_engine_kwargs(),
    )


def _map_phoenix(row: dict[str, Any]) -> PositionMapping:
    structure_type = text_value(row.get(PositionCol.STRUCTURE_TYPE))
    return _supported_mapping(
        row,
        product_type="PhoenixOption",
        product_kwargs={
            **_autocallable_core_kwargs(row),
            "barrier_config": _phoenix_barrier_config(row),
            "coupon_config": _phoenix_coupon_config(row),
            "payoff_config": _autocallable_payoff_config(row, partial_protection=PARTIAL_PROTECTION_TAG in structure_type),
            "accrual_config": _autocallable_accrual_config(row),
            "_otc_ki_observation_convention": _ki_observation_convention(row),
            "_otc_lifecycle_knocked_in": _autocallable_knocked_in(row),
            "_otc_lifecycle_state": text_value(row.get(PositionCol.TRADE_STATUS)),
        },
        engine_name="PhoenixQuadEngine",
        engine_kwargs=_quad_engine_kwargs(),
    )


def _supported_mapping(
    row: dict[str, Any],
    *,
    product_type: str,
    product_kwargs: dict[str, Any],
    engine_name: str,
    engine_kwargs: dict[str, Any] | None = None,
) -> PositionMapping:
    return PositionMapping(
        underlying=normalize_symbol(row.get(PositionCol.UNDERLYING_CODE)),
        product_type=product_type,
        product_kwargs=make_json_safe(product_kwargs),
        engine_name=engine_name,
        engine_kwargs=engine_kwargs or {},
        quantity=_signed_quantity(row),
        entry_price=0.0,
        status=_position_status(row),
        mapping_status=SUPPORTED_STATUS,
    )


def _unsupported_mapping(row: dict[str, Any], reason: str) -> PositionMapping:
    return PositionMapping(
        underlying=normalize_symbol(row.get(PositionCol.UNDERLYING_CODE)),
        product_type=text_value(row.get(PositionCol.STRUCTURE_TYPE)) or "Unsupported",
        product_kwargs={},
        engine_name="Unsupported",
        engine_kwargs={},
        quantity=_signed_quantity(row),
        entry_price=0.0,
        status=_position_status(row),
        mapping_status=UNSUPPORTED_STATUS,
        mapping_error=reason,
    )


def _error_mapping(row: dict[str, Any], reason: str) -> PositionMapping:
    mapping = _unsupported_mapping(row, reason)
    return PositionMapping(
        **{**mapping.__dict__, "mapping_status": ERROR_STATUS, "mapping_error": reason}
    )


def _signed_quantity(row: dict[str, Any]) -> float:
    side = text_value(row.get(PositionCol.DIRECTION))
    sign = -1.0 if side == Direction.SELL else 1.0
    return sign


def _position_status(row: dict[str, Any]) -> str:
    trade_state = text_value(row.get(PositionCol.TRADE_STATUS))
    return "closed" if trade_state in CLOSED_STATUSES else "open"


def _contract_multiplier(row: dict[str, Any]) -> float:
    notional = parse_number(row.get(PositionCol.NOTIONAL), None) or parse_number(row.get(PositionCol.INITIAL_NOTIONAL), None)
    initial_price = _required_number(row, PositionCol.INITIAL_PRICE)
    if notional is None or notional <= 0:
        raise ValueError("Notional is required and must be positive")
    return notional / initial_price


def _required_number(row: dict[str, Any], key: str) -> float:
    value = parse_number(row.get(key), None)
    if value is None:
        raise ValueError(f"{key} is required")
    return value


def _option_type(row: dict[str, Any]) -> str:
    option_type = text_value(row.get(PositionCol.OPTION_TYPE))
    if option_type == OptionType.PUT:
        return "PUT"
    return "CALL"


def _date_iso(row: dict[str, Any], key: str) -> str | None:
    value = parse_date(row.get(key))
    return value.date().isoformat() if value is not None else None


def _vanilla_date_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    kwargs = {
        "exercise_date": _date_iso(row, PositionCol.FINAL_OBSERVATION_DATE) or _date_iso(row, PositionCol.MATURITY_DATE),
        "settlement_date": _date_iso(row, PositionCol.SETTLEMENT_DATE),
        "contract_multiplier": _contract_multiplier(row),
    }
    return {key: value for key, value in kwargs.items() if value is not None}


def _sharkfin_core_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    initial_price = _required_number(row, PositionCol.INITIAL_PRICE)
    knock_out_rate = _list_value(parse_number_list(row.get(PositionCol.KNOCK_OUT_COUPON)), 0)
    no_hit_rate = (
        parse_number(row.get(PositionCol.COUPON_RATE), None)
        or parse_number(row.get(PositionCol.DIVIDEND_COUPON), None)
        or parse_number(row.get(PositionCol.NO_KNOCK_IN_COUPON), 0.0)
        or 0.0
    )
    kwargs = {
        "strike": _required_number(row, PositionCol.STRIKE_PRICE),
        "option_type": _option_type(row),
        "participation_rate": parse_number(row.get(PositionCol.PARTICIPATION_RATE), 1.0) or 1.0,
        "knock_out_rebate": initial_price * (knock_out_rate or 0.0),
        "no_hit_rebate": initial_price * no_hit_rate,
        "pay_at_hit": False,
        "exercise_date": _date_iso(row, PositionCol.FINAL_OBSERVATION_DATE) or _date_iso(row, PositionCol.MATURITY_DATE),
        "settlement_date": _date_iso(row, PositionCol.SETTLEMENT_DATE),
        "contract_multiplier": _contract_multiplier(row),
    }
    return {key: value for key, value in kwargs.items() if value is not None}


def _single_sharkfin_barrier(row: dict[str, Any]) -> float:
    barriers = parse_number_list(row.get(PositionCol.KNOCK_OUT_BARRIER))
    if not barriers:
        return _required_number(row, PositionCol.KNOCK_OUT_BARRIER)
    strike = _required_number(row, PositionCol.STRIKE_PRICE)
    if _option_type(row) == "PUT":
        below_strike = [barrier for barrier in barriers if barrier < strike]
        return below_strike[0] if below_strike else min(barriers)
    above_strike = [barrier for barrier in barriers if barrier > strike]
    return above_strike[0] if above_strike else max(barriers)


def _double_sharkfin_barriers(row: dict[str, Any]) -> tuple[float, float]:
    barriers = [
        *parse_number_list(row.get(PositionCol.KNOCK_OUT_BARRIER)),
        *parse_number_list(row.get(PositionCol.KNOCK_OUT_BARRIER_2)),
        *parse_number_list(row.get(PositionCol.KNOCK_IN_BARRIER)),
    ]
    distinct = sorted({barrier for barrier in barriers if barrier > 0.0})
    if len(distinct) < 2:
        raise ValueError(
            "Double Sharkfin requires two barrier levels from "
            "Knock-Out Barrier / Knock-Out Barrier 2 / Knock-In Barrier"
        )
    return distinct[0], distinct[-1]


def _sharkfin_observation_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    observation_dates = parse_date_list(row.get(PositionCol.KNOCK_OUT_OBSERVATION_DATES))
    if not observation_dates:
        return {"observation_type": "CONTINUOUS"}
    barrier = _single_sharkfin_barrier(row)
    return {
        "observation_type": "DISCRETE",
        "observation_schedule": _single_sharkfin_schedule(
            dates=observation_dates,
            barrier=barrier,
            payoff=_sharkfin_core_kwargs(row)["knock_out_rebate"],
        ),
    }


def _double_sharkfin_observation_kwargs(
    row: dict[str, Any],
    lower_barrier: float,
    upper_barrier: float,
) -> dict[str, Any]:
    observation_dates = parse_date_list(row.get(PositionCol.KNOCK_OUT_OBSERVATION_DATES))
    if not observation_dates:
        return {"observation_type": "CONTINUOUS"}
    return {
        "observation_type": "DISCRETE",
        "observation_schedule": _double_sharkfin_schedule(
            dates=observation_dates,
            lower_barrier=lower_barrier,
            upper_barrier=upper_barrier,
            payoff=_sharkfin_core_kwargs(row)["knock_out_rebate"],
        ),
    }


def _autocallable_core_kwargs(row: dict[str, Any]) -> dict[str, Any]:
    kwargs = {
        "initial_price": _required_number(row, PositionCol.INITIAL_PRICE),
        "strike": _required_number(row, PositionCol.STRIKE_PRICE),
        "initial_date": _date_iso(row, PositionCol.START_DATE),
        "exercise_date": _date_iso(row, PositionCol.FINAL_OBSERVATION_DATE) or _date_iso(row, PositionCol.MATURITY_DATE),
        "settlement_date": _date_iso(row, PositionCol.SETTLEMENT_DATE),
        "contract_multiplier": _contract_multiplier(row),
        "is_reverse": False,
    }
    return {key: value for key, value in kwargs.items() if value is not None}


def _snowball_barrier_config(row: dict[str, Any]) -> dict[str, Any]:
    ko_dates = parse_date_list(row.get(PositionCol.KNOCK_OUT_OBSERVATION_DATES))
    ko_barriers = parse_number_list(row.get(PositionCol.KNOCK_OUT_BARRIER))
    ko_rates = parse_number_list(row.get(PositionCol.KNOCK_OUT_COUPON))
    ki_config = _ki_barrier_config(row)
    config: dict[str, Any] = {
        "ko_barrier": _scalar_or_list(ko_barriers) or _required_number(row, PositionCol.KNOCK_OUT_BARRIER),
        "ko_rate": _scalar_or_list(ko_rates) or (parse_number(row.get(PositionCol.DIVIDEND_COUPON), 0.0) or 0.0),
        "ko_observation_type": "DISCRETE",
        **ki_config,
    }
    if ko_dates:
        config["ko_observation_schedule"] = _single_barrier_schedule(
            dates=ko_dates,
            barriers=ko_barriers,
            rates=ko_rates,
            annualized=_is_annualized(row),
        )
    return config


def _phoenix_barrier_config(row: dict[str, Any]) -> dict[str, Any]:
    ko_dates = parse_date_list(row.get(PositionCol.KO_COUPON_OBSERVATION_DATES)) or parse_date_list(row.get(PositionCol.KNOCK_OUT_OBSERVATION_DATES))
    ko_barriers = parse_number_list(row.get(PositionCol.KNOCK_OUT_BARRIER))
    ko_rates = parse_number_list(row.get(PositionCol.KNOCK_OUT_COUPON))
    ki_config = _ki_barrier_config(row)
    config: dict[str, Any] = {
        "ko_barrier": _scalar_or_list(ko_barriers) or _required_number(row, PositionCol.KNOCK_OUT_BARRIER),
        "ko_rate": _scalar_or_list(ko_rates) or 0.0,
        "ko_observation_type": "DISCRETE",
        **ki_config,
    }
    if ko_dates:
        config["ko_observation_schedule"] = _single_barrier_schedule(
            dates=ko_dates,
            barriers=ko_barriers,
            rates=ko_rates or [0.0],
            annualized=_is_annualized(row),
        )
    return config


def _phoenix_coupon_config(row: dict[str, Any]) -> dict[str, Any]:
    coupon_barriers = parse_number_list(row.get(PositionCol.COUPON_BARRIER))
    coupon_rates = parse_number_list(row.get(PositionCol.COUPON_BARRIER_RATE))
    return {
        "coupon_barrier": _scalar_or_list(coupon_barriers) or _required_number(row, PositionCol.COUPON_BARRIER),
        "coupon_rate": coupon_rates[0] if coupon_rates else 0.0,
        "coupon_pay_type": "INSTANT",
        "memory_coupon": False,
    }


def _autocallable_payoff_config(row: dict[str, Any], *, partial_protection: bool) -> dict[str, Any]:
    config = {
        "rebate_rate": parse_number(row.get(PositionCol.DIVIDEND_COUPON), 0.0) or parse_number(row.get(PositionCol.NO_KNOCK_IN_COUPON), 0.0) or 0.0,
        "include_principal": False,
        "participation_rate": parse_number(row.get(PositionCol.PARTICIPATION_RATE), 1.0) or 1.0,
        "protection_type": "NONE",
        "protection_rate": 0.0,
    }
    if partial_protection:
        config["protection_type"] = "PARTIAL"
        config["protection_rate"] = parse_number(row.get(PositionCol.KI_MIN_RETURN_RATE), 0.5) or 0.5
    return config


def _autocallable_accrual_config(row: dict[str, Any]) -> dict[str, Any]:
    annualized = _is_annualized(row)
    config: dict[str, Any] = {
        "coupon_pay_type": "INSTANT",
        "is_annualized": annualized,
        "is_annualized_ko": annualized,
        "is_annualized_ki": _is_ki_annualized(row),
        "is_annualized_rebate": annualized,
    }
    accrual_factors = _autocallable_accrual_factors(row)
    if accrual_factors:
        config["accrual_factors"] = accrual_factors
    return config


def _is_annualized(row: dict[str, Any]) -> bool:
    return text_value(row.get(PositionCol.ANNUALIZED)) == YesNo.YES


def _is_ki_annualized(row: dict[str, Any]) -> bool:
    return text_value(row.get(PositionCol.KI_ANNUALIZED)) == YesNo.YES


def _autocallable_accrual_factors(row: dict[str, Any]) -> list[float]:
    if PHOENIX_TAG in text_value(row.get(PositionCol.STRUCTURE_TYPE)):
        return parse_accrual_factor_list(row.get(PositionCol.DAY_COUNT_FACTORS))
    return parse_accrual_day_count_list(row.get(PositionCol.KNOCK_OUT_DAY_COUNTS))


def parse_accrual_day_count_list(value: Any, *, basis: float = 365.0) -> list[float]:
    return [
        day_count / basis
        for day_count in parse_accrual_factor_list(value)
    ]


def parse_accrual_factor_list(value: Any) -> list[float]:
    if value in (None, ""):
        return []
    if isinstance(value, (int, float)):
        return [float(value)]
    factors = []
    for part in _split_list_text(value):
        factor = _parse_accrual_factor(part)
        if factor is not None:
            factors.append(factor)
    return factors


def _split_list_text(value: Any) -> list[str]:
    text = str(value).strip()
    if not text:
        return []
    if text.startswith("[") and text.endswith("]"):
        text = text[1:-1]
    return [part.strip() for part in text.replace("，", ",").split(",") if part.strip()]


def _parse_accrual_factor(value: str) -> float | None:
    if "/" not in value:
        return parse_number(value, None)
    numerator_text, denominator_text = (part.strip() for part in value.split("/", 1))
    numerator = parse_number(numerator_text, None)
    denominator = parse_number(denominator_text, None)
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator


def _ki_barrier_config(row: dict[str, Any]) -> dict[str, Any]:
    convention = _ki_observation_convention(row)
    config: dict[str, Any] = {
        "ki_barrier": None if convention == "NONE" else parse_number(row.get(PositionCol.KNOCK_IN_BARRIER), None),
        "ki_observation_type": "DISCRETE",
        "ki_continuous": False,
    }
    if convention == "DAILY":
        schedule = _daily_ki_schedule(row, config["ki_barrier"])
    elif convention == "EUROPEAN":
        schedule = _european_ki_schedule(row, config["ki_barrier"])
    else:
        schedule = None
    if schedule is not None:
        config["ki_observation_schedule"] = schedule
    return config


def _ki_observation_convention(row: dict[str, Any]) -> str:
    custom_structure = text_value(row.get(PositionCol.CUSTOM_STRUCTURE))
    if CUSTOM_NO_KNOCK_IN in custom_structure:
        return "NONE"
    if CUSTOM_EUROPEAN in custom_structure:
        return "EUROPEAN"
    return "DAILY"


def _has_no_knock_in_terms(row: dict[str, Any]) -> bool:
    return CUSTOM_NO_KNOCK_IN in text_value(row.get(PositionCol.CUSTOM_STRUCTURE))


def _autocallable_knocked_in(row: dict[str, Any]) -> bool:
    return (
        _has_no_knock_in_terms(row)
        or text_value(row.get(PositionCol.ALREADY_KNOCKED_IN)) == YesNo.YES
        or text_value(row.get(PositionCol.TRADE_STATUS)) == TradeStatus.KNOCKED_IN
    )


def _daily_ki_schedule(row: dict[str, Any], barrier: Any) -> dict[str, Any] | None:
    start_date = parse_date(row.get(PositionCol.START_DATE))
    exercise_date = _autocallable_exercise_date(row)
    if barrier is None or start_date is None or exercise_date is None:
        return None
    dates = []
    current = start_date.date() + timedelta(days=1)
    end = exercise_date.date()
    while current <= end:
        if _is_china_sse_business_day(current):
            dates.append(datetime.combine(current, datetime.min.time()))
        current += timedelta(days=1)
    return _ki_schedule(dates=dates, barrier=barrier, frequency="DAILY")


def _european_ki_schedule(row: dict[str, Any], barrier: Any) -> dict[str, Any] | None:
    exercise_date = _autocallable_exercise_date(row)
    if barrier is None or exercise_date is None:
        return None
    return _ki_schedule(dates=[exercise_date], barrier=barrier, frequency="CUSTOM")


def _autocallable_exercise_date(row: dict[str, Any]) -> datetime | None:
    return parse_date(row.get(PositionCol.FINAL_OBSERVATION_DATE)) or parse_date(row.get(PositionCol.MATURITY_DATE))


def _ki_schedule(*, dates: list[datetime], barrier: Any, frequency: str) -> dict[str, Any] | None:
    if not dates:
        return None
    return {
        "records": [
            {"observation_date": date.date().isoformat(), "barrier": barrier}
            for date in dates
        ],
        "aggregation_mode": "STOP_FIRST_HIT",
        "frequency": frequency,
    }


def _is_china_sse_business_day(day: date) -> bool:
    return _shared_is_business_day(day)


def _single_barrier_schedule(
    *,
    dates: list[datetime],
    barriers: list[float],
    rates: list[float],
    annualized: bool,
) -> dict[str, Any]:
    records = []
    for index, date in enumerate(dates):
        barrier = _list_value(barriers, index)
        record: dict[str, Any] = {
            "observation_date": date.date().isoformat(),
            "barrier": barrier,
        }
        rate = _list_value(rates, index)
        if rate is not None:
            record["return_rate"] = rate
            record["is_rate_annualized"] = annualized
        records.append(record)
    return {"records": records, "aggregation_mode": "STOP_FIRST_HIT", "frequency": "CUSTOM"}


def _single_sharkfin_schedule(
    *,
    dates: list[datetime],
    barrier: float,
    payoff: float,
) -> dict[str, Any]:
    return {
        "records": [
            {
                "observation_date": date.date().isoformat(),
                "barrier": barrier,
                "payoff": payoff,
            }
            for date in dates
        ],
        "aggregation_mode": "STOP_FIRST_HIT",
        "frequency": "CUSTOM",
    }


def _double_sharkfin_schedule(
    *,
    dates: list[datetime],
    lower_barrier: float,
    upper_barrier: float,
    payoff: float,
) -> dict[str, Any]:
    return {
        "records": [
            {
                "observation_date": date.date().isoformat(),
                "lower_barrier": lower_barrier,
                "upper_barrier": upper_barrier,
                "payoff": payoff,
            }
            for date in dates
        ],
        "aggregation_mode": "STOP_FIRST_HIT",
        "frequency": "CUSTOM",
    }


def _list_value(values: list[float], index: int) -> float | None:
    if not values:
        return None
    if index < len(values):
        return values[index]
    return values[-1]


def _scalar_or_list(values: list[float]) -> float | list[float] | None:
    if not values:
        return None
    if len(values) == 1:
        return values[0]
    return values


def _quad_engine_kwargs() -> dict[str, Any]:
    # Intentionally no params_kwargs. The pricer's adaptive chain picks a grid at
    # call time; baking grid_points here would suppress that escalation.
    return {"params_type": "quad_params"}
