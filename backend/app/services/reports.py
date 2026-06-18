from __future__ import annotations

import re
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from openpyxl import Workbook
from sqlalchemy.orm import Session, sessionmaker

from .. import database
from ..config import Settings
from ..models import (
    RFQ,
    Portfolio,
    PricingParameterProfile,
    ReportJob,
    ReportStatus,
    TaskKind,
    TaskRun,
    TaskStatus,
)
from ..schemas import ReportJobCreate
from .quantark import calculate_portfolio_risk
from .portfolio_membership import resolve_positions
from .risk_engine import pricing_position_markets
from .task_runner import mark_task_finished, mark_task_running, update_task_progress


def create_report_job(
    session: Session, settings: Settings, request: ReportJobCreate
) -> ReportJob:
    settings.artifact_dir.mkdir(parents=True, exist_ok=True)
    job = ReportJob(
        report_type=request.report_type,
        status=ReportStatus.RUNNING.value,
        request_payload=request.model_dump(mode="json"),
        result_payload={},
        artifact_paths={},
    )
    session.add(job)
    session.flush()
    _complete_report_job(session, settings, job, request)
    return job


def queue_report_job(
    session: Session,
    request: ReportJobCreate,
) -> tuple[ReportJob, TaskRun]:
    job = ReportJob(
        report_type=request.report_type,
        status=ReportStatus.QUEUED.value,
        request_payload=request.model_dump(mode="json"),
        result_payload={},
        artifact_paths={},
    )
    session.add(job)
    session.flush()
    task = TaskRun(
        kind=TaskKind.REPORT_JOB.value,
        status=TaskStatus.QUEUED.value,
        portfolio_id=request.portfolio_id,
        report_job_id=job.id,
        progress_current=0,
        progress_total=3,
        message="Queued report job",
    )
    session.add(task)
    session.flush()
    return job, task


def execute_report_job_task(
    task_id: int,
    report_job_id: int,
    settings: Settings,
    session_factory: sessionmaker | None = None,
) -> None:
    session = (session_factory or database.SessionLocal)()
    try:
        _execute_report_job_task(session, task_id, report_job_id, settings)
    finally:
        session.close()


def _execute_report_job_task(
    session: Session,
    task_id: int,
    report_job_id: int,
    settings: Settings,
) -> None:
    try:
        job = session.get(ReportJob, report_job_id)
        if job is None:
            mark_task_finished(
                session,
                task_id,
                status=TaskStatus.FAILED.value,
                error=f"Report job not found: {report_job_id}",
            )
            session.commit()
            return
        request = ReportJobCreate.model_validate(job.request_payload)
        mark_task_running(session, task_id, message="Building report payload", total=3)
        session.commit()
        _complete_report_job(
            session,
            settings,
            job,
            request,
            task_id=task_id,
        )
        status = job.status
        mark_task_finished(
            session,
            task_id,
            status=status,
            message=(
                "Report generated"
                if status == TaskStatus.COMPLETED.value
                else "Report generated with issues"
            ),
        )
        session.commit()
    except Exception as exc:
        session.rollback()
        job = session.get(ReportJob, report_job_id)
        if job is not None:
            job.status = ReportStatus.FAILED.value
        mark_task_finished(
            session,
            task_id,
            status=TaskStatus.FAILED.value,
            message="Report generation failed",
            error=str(exc),
        )
        session.commit()


def _complete_report_job(
    session: Session,
    settings: Settings,
    job: ReportJob,
    request: ReportJobCreate,
    *,
    task_id: int | None = None,
) -> None:
    settings.artifact_dir.mkdir(parents=True, exist_ok=True)
    job.status = ReportStatus.RUNNING.value
    if task_id is not None:
        update_task_progress(
            session, task_id, current=0, total=3, message="Building report payload"
        )

    result = _build_report_payload(session, request)
    if task_id is not None:
        update_task_progress(
            session, task_id, current=1, total=3, message="Writing HTML report"
        )
    artifact_stem = _report_artifact_stem(job, request)
    html_path = settings.artifact_dir / f"{artifact_stem}.html"
    xlsx_path = settings.artifact_dir / f"{artifact_stem}.xlsx"
    _write_html(html_path, request.title, result)
    if task_id is not None:
        update_task_progress(
            session, task_id, current=2, total=3, message="Writing Excel report"
        )
    _write_xlsx(xlsx_path, request.title, result)
    job.result_payload = result
    job.artifact_paths = {"html": str(html_path), "excel": str(xlsx_path)}
    job.status = _report_status_from_payload(result)
    if task_id is not None:
        update_task_progress(
            session, task_id, current=3, total=3, message="Report generated"
        )
    session.flush()


def _report_artifact_stem(job: ReportJob, request: ReportJobCreate) -> str:
    task_name = _safe_artifact_name(request.title or request.report_type)
    timestamp = job.created_at.strftime("%Y%m%d_%H%M%S_%f")
    return f"{task_name}_{timestamp}"


def _safe_artifact_name(value: str) -> str:
    name = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
    return name or "report"


def _build_report_payload(session: Session, request: ReportJobCreate) -> dict[str, Any]:
    if request.report_type == "rfq" and request.rfq_id is not None:
        rfq = session.get(RFQ, request.rfq_id)
        if not rfq:
            return {"error": "RFQ not found"}
        return {
            "rfq": rfq.quote_payload,
            "status": rfq.status,
            "request": rfq.request_payload,
        }
    portfolio = (
        session.get(Portfolio, request.portfolio_id)
        if request.portfolio_id
        else session.query(Portfolio).first()
    )
    if not portfolio:
        return {"summary": {"message": "No portfolio available"}, "positions": []}
    resolved = resolve_positions(portfolio, session)
    portfolio_like = SimpleNamespace(
        id=portfolio.id,
        name=portfolio.name,
        base_currency=portfolio.base_currency,
        positions=resolved,
    )
    risk = calculate_portfolio_risk(
        portfolio_like,  # type: ignore[arg-type]
        position_markets=pricing_position_markets(
            session,
            resolved,
            pricing_parameter_profile_id=request.pricing_parameter_profile_id,
        ),
    )
    return {
        "portfolio": {"id": portfolio.id, "name": portfolio.name},
        "pricing_parameter_profile": _pricing_profile_payload(
            session, request.pricing_parameter_profile_id
        ),
        "risk": risk,
    }


def _pricing_profile_payload(
    session: Session,
    pricing_parameter_profile_id: int | None,
) -> dict[str, Any] | None:
    if pricing_parameter_profile_id is None:
        return None
    profile = session.get(PricingParameterProfile, pricing_parameter_profile_id)
    if profile is None:
        return {"id": pricing_parameter_profile_id, "missing": True}
    return {
        "id": profile.id,
        "name": profile.name,
        "valuation_date": profile.valuation_date.isoformat(),
        "source_type": profile.source_type,
        "source_path": profile.source_path,
    }


def _report_status_from_payload(payload: dict[str, Any]) -> str:
    if payload.get("error"):
        return ReportStatus.COMPLETED_WITH_ERRORS.value
    for row in payload.get("risk", {}).get("positions", []) or []:
        if not row.get("pricing_ok") or not row.get("greeks_ok"):
            return ReportStatus.COMPLETED_WITH_ERRORS.value
    return ReportStatus.COMPLETED.value


_MONEY_DISPLAY = [
    ("market_value", "Market value"), ("pnl", "PnL"),
    ("gross_notional", "Gross notional"), ("one_day_var_proxy", "1D VaR proxy"),
    ("vega", "Vega"), ("theta", "Theta"), ("rho", "Rho"), ("rho_q", "Rho (q)"),
    ("delta_cash", "Delta cash"), ("gamma_cash", "Gamma cash"),
    ("position_count", "Positions"),
]
_SHARED_DISPLAY = [("delta", "Delta"), ("gamma", "Gamma"), ("delta_proxy", "Delta proxy")]


def _metric_card(label: str, value: Any) -> str:
    try:
        rendered = f"{float(value):.4f}"
    except (TypeError, ValueError):
        rendered = "0.0000"
    return f'<div class="metric"><div>{label}</div><div class="value">{rendered}</div></div>'


def _write_html(path: Path, title: str, payload: dict[str, Any]) -> None:
    risk = payload.get("risk", {}) or {}
    totals = risk.get("totals") or {}
    by_currency = risk.get("by_currency") or {}
    shared = risk.get("shared") or {}
    positions = risk.get("positions", [])

    if totals:
        top_block = '<div class="grid">' + "".join(
            _metric_card(label, totals.get(key, 0)) for key, label in _MONEY_DISPLAY[:4]
        ) + "</div>"
    else:
        top_block = '<p class="note">Mixed currency — see the per-currency breakdown below.</p>'

    by_currency_html = ""
    if by_currency:
        sections = []
        for ccy in sorted(by_currency):
            bucket = by_currency[ccy]
            cards = "".join(_metric_card(label, bucket.get(key, 0)) for key, label in _MONEY_DISPLAY)
            sections.append(f'<h2>{ccy}</h2><div class="grid">{cards}</div>')
        by_currency_html = "<h2>By currency</h2>" + "".join(sections)

    shared_html = ""
    if shared:
        cards = "".join(_metric_card(label, shared.get(key, 0)) for key, label in _SHARED_DISPLAY)
        shared_html = f'<h2>Shared (currency-invariant)</h2><div class="grid">{cards}</div>'

    rows = "\n".join(
        f"<tr><td>{p.get('position_id')}</td><td>{p.get('underlying')}</td><td>{p.get('product_type')}</td>"
        f"<td>{p.get('quantity')}</td><td>{p.get('market_value'):.4f}</td><td>{p.get('pnl'):.4f}</td></tr>"
        for p in positions
    )
    path.write_text(
        f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8" />
  <title>{title}</title>
  <style>
    body {{ font-family: Inter, Arial, sans-serif; margin: 32px; color: #172033; }}
    table {{ border-collapse: collapse; width: 100%; margin-top: 24px; }}
    th, td {{ border-bottom: 1px solid #d7deea; padding: 10px; text-align: left; }}
    .grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-top: 8px; }}
    .metric {{ border: 1px solid #d7deea; border-radius: 8px; padding: 16px; }}
    .value {{ font-size: 24px; font-weight: 700; }}
    .note {{ color: #6b7686; font-style: italic; }}
    h2 {{ margin-top: 28px; font-size: 16px; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  {top_block}
  {by_currency_html}
  {shared_html}
  <table>
    <thead><tr><th>ID</th><th>Underlying</th><th>Product</th><th>Qty</th><th>MV</th><th>PnL</th></tr></thead>
    <tbody>{rows}</tbody>
  </table>
</body>
</html>
""",
        encoding="utf-8",
    )


def _write_xlsx(path: Path, title: str, payload: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([title])
    totals = payload.get("risk", {}).get("totals") or {}
    for key, value in totals.items():
        ws.append([key, value])

    risk = payload.get("risk", {}) or {}
    by_currency = risk.get("by_currency") or {}
    shared = risk.get("shared") or {}
    if by_currency or shared:
        ccy_ws = wb.create_sheet("By Currency")
        ccy_ws.append(["currency", "metric", "value"])
        for ccy in sorted(by_currency):
            bucket = by_currency[ccy]
            for key, _label in _MONEY_DISPLAY:
                if key in bucket:
                    ccy_ws.append([ccy, key, bucket[key]])
        for key, _label in _SHARED_DISPLAY:
            if key in shared:
                ccy_ws.append(["(shared)", key, shared[key]])

    positions_ws = wb.create_sheet("Positions")
    positions_ws.append(
        [
            "position_id",
            "underlying",
            "product_type",
            "quantity",
            "price",
            "market_value",
            "pnl",
        ]
    )
    for row in payload.get("risk", {}).get("positions", []):
        positions_ws.append(
            [
                row.get("position_id"),
                row.get("underlying"),
                row.get("product_type"),
                row.get("quantity"),
                row.get("price"),
                row.get("market_value"),
                row.get("pnl"),
            ]
        )
    wb.save(path)
