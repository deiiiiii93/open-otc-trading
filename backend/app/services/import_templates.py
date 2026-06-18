"""Generate the downloadable English import templates.

Each template is a two-sheet workbook driven entirely by the ``ColumnSpec`` lists
in :mod:`app.services.import_schema`:

1. A **data sheet** — a bold header row followed by a couple of illustrative
   example rows, so the user can see how list-style cells are formatted.
2. An **Instructions sheet** — one row per column documenting whether it is
   required, its type, its allowed enum values, and a usage note.

Because the columns come from the same module the parsers read, the blank
template can never list a column the adapter ignores (or omit one it needs).
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .import_schema import (
    POSITION_COLUMNS,
    POSITIONS_SHEET_NAME,
    PRICING_COLUMNS,
    PRICING_SHEET_NAME,
    ColumnSpec,
)

INSTRUCTIONS_SHEET_NAME = "Instructions"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_REQUIRED_FILL = PatternFill("solid", fgColor="FCE4D6")
_WRAP = Alignment(vertical="top", wrap_text=True)


def _example_count(columns: list[ColumnSpec]) -> int:
    return max((len(spec.examples) for spec in columns), default=0)


def _style_header_row(worksheet: Worksheet, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    worksheet.freeze_panes = "A2"


def _build_data_sheet(workbook: Workbook, title: str, columns: list[ColumnSpec]) -> Worksheet:
    worksheet = workbook.active
    assert worksheet is not None  # a fresh Workbook always has an active sheet
    worksheet.title = title
    worksheet.append([spec.name for spec in columns])
    for index in range(_example_count(columns)):
        worksheet.append([
            spec.examples[index] if index < len(spec.examples) else None
            for spec in columns
        ])
    _style_header_row(worksheet, len(columns))
    for col, spec in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(col)].width = max(
            14, min(32, len(spec.name) + 4)
        )
    return worksheet


def _build_instructions_sheet(workbook: Workbook, columns: list[ColumnSpec]) -> Worksheet:
    worksheet = workbook.create_sheet(INSTRUCTIONS_SHEET_NAME)
    headers = ["Column", "Required", "Type", "Allowed Values", "Notes"]
    worksheet.append(headers)
    for spec in columns:
        worksheet.append([
            spec.name,
            "Required" if spec.required else "Optional",
            spec.dtype,
            ", ".join(spec.allowed),
            spec.notes,
        ])
        if spec.required:
            worksheet.cell(row=worksheet.max_row, column=2).fill = _REQUIRED_FILL
    _style_header_row(worksheet, len(headers))
    widths = [30, 10, 12, 46, 60]
    for col, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(col)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP
    return worksheet


def _build_template(sheet_title: str, columns: list[ColumnSpec]) -> Workbook:
    workbook = Workbook()
    _build_data_sheet(workbook, sheet_title, columns)
    _build_instructions_sheet(workbook, columns)
    return workbook


def build_positions_template() -> Workbook:
    return _build_template(POSITIONS_SHEET_NAME, POSITION_COLUMNS)


def build_pricing_parameters_template() -> Workbook:
    return _build_template(PRICING_SHEET_NAME, PRICING_COLUMNS)


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def positions_template_bytes() -> bytes:
    return workbook_to_bytes(build_positions_template())


def pricing_parameters_template_bytes() -> bytes:
    return workbook_to_bytes(build_pricing_parameters_template())
